"""기초율표(경험률) 파싱 · 정규화.

고객/계리사의 PUC 워크북 '기초율표' 시트 또는 보험개발원 기초율 파일을 읽어,
연령별 표준 기초율(퇴직율·사망률 남/여·승급률)을 정규화 구조로 추출한다.
개발원 기초율은 3년 주기로 변경되고, 300인 미만/이상(사업장 규모)별로 퇴직률·
승급률이 다르므로 밴드별로 함께 보존한다. 과거 산출에 적용된 기초율은 감사추적을
위해 세트(버전) 단위로 보존해야 하므로 여기서는 '한 세트'를 만드는 파싱만 담당한다.

지원 레이아웃 두 가지:
  (A) 표준(연령 컬럼 있음): 헤더행(연령·퇴직률·사망률…)을 인식해 열 매핑.
      300인 밴드 컬럼(퇴직률(300인미만)/(300인이상), 승급률(미적용)/(300인미만)/(300인이상))도
      헤더로 인식하면 밴드별로 저장.
  (B) 개발원(연령 컬럼 없음): 상단에 '남자/여자', '300인 미만/이상', '미적용' 소제목이
      있고 데이터는 시작연령(기본 15세)부터 행 순서로 나열 → 위치로 파싱.
"""

from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Dict, List, Optional, Union

# 정규화 필드(연령 행 단위) — 표준 단일 밴드
RATE_FIELDS = ["age", "withdrawal", "mort_m", "mort_f", "raise_rate"]
# 300인 밴드별 확장 필드
BAND_FIELDS = ["withdrawal_lt300", "withdrawal_ge300",
               "raise_none", "raise_lt300", "raise_ge300"]
# 표시용 한글 라벨
RATE_LABELS = {
    "age": "연령", "withdrawal": "퇴직율", "mort_m": "사망률(남)",
    "mort_f": "사망률(여)", "raise_rate": "승급률",
    "withdrawal_lt300": "퇴직률(300인미만)", "withdrawal_ge300": "퇴직률(300인이상)",
    "raise_none": "승급률(미적용)", "raise_lt300": "승급률(300인미만)",
    "raise_ge300": "승급률(300인이상)",
}

_SHEET_CANDIDATES = ["기초율표", "기초율", "기초율표(경험률)"]


def _num(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# 컬럼 별칭(헤더 인식) — 어떤 양식으로 올려도 헤더로 맞춘다.
_COL_ALIASES = {
    "age": ["연령", "나이", "도달연령", "age"],
    "withdrawal": ["퇴직률", "퇴직율", "이직률", "이직율", "퇴직", "withdrawal"],
    "mort_m": ["사망률(남)", "사망률남", "남자사망률", "남사망률", "male_qx", "남성사망률"],
    "mort_f": ["사망률(여)", "사망률여", "여자사망률", "여사망률", "female_qx", "여성사망률"],
    "raise_rate": ["승급률", "승급율", "임금상승률", "임금인상률", "호봉상승률", "raise"],
    # 300인 밴드별 컬럼(표준 양식에서 명시 헤더로 올릴 때)
    "withdrawal_lt300": ["퇴직률(300인미만)", "퇴직율(300인미만)", "퇴직률300인미만",
                         "퇴직률(300미만)", "퇴직률300미만", "withdrawal_lt300"],
    "withdrawal_ge300": ["퇴직률(300인이상)", "퇴직율(300인이상)", "퇴직률300인이상",
                         "퇴직률(300이상)", "퇴직률300이상", "withdrawal_ge300"],
    "raise_none": ["승급률(미적용)", "승급율(미적용)", "승급률미적용", "raise_none"],
    "raise_lt300": ["승급률(300인미만)", "승급율(300인미만)", "승급률300인미만",
                    "승급률(300미만)", "raise_lt300"],
    "raise_ge300": ["승급률(300인이상)", "승급율(300인이상)", "승급률300인이상",
                    "승급률(300이상)", "raise_ge300"],
}


def _bn(s) -> str:
    """기초율 헤더 정규화 — 공백만 제거(괄호는 유지: 사망률(남)/(여) 구분)."""
    return "".join(str(s).split()).lower()


def _detect_header(grid):
    """헤더 행을 찾아 {필드: 열index} 매핑을 반환. 못 찾으면 (None, {})."""
    alias_norm = {f: [_bn(a) for a in al] for f, al in _COL_ALIASES.items()}
    for ridx, row in enumerate(grid[:15]):
        norm = {}
        for cidx, cell in enumerate(row):
            if cell is None:
                continue
            norm.setdefault(_bn(cell), cidx)
        colmap = {}
        for field, na in alias_norm.items():
            for a in na:
                if a in norm:
                    colmap[field] = norm[a]
                    break
        # 밴드 헤더가 잡히면 일반 '퇴직률/승급률'과 겹칠 수 있으므로 밴드 우선.
        # 연령 + (퇴직률 또는 사망률) 이 잡히면 헤더 행으로 인정
        wanted_w = colmap.keys() & {"withdrawal", "withdrawal_lt300", "withdrawal_ge300",
                                    "mort_m", "mort_f"}
        if "age" in colmap and wanted_w:
            return ridx, colmap
    return None, {}


def _detect_dev_layout(grid):
    """개발원 레이아웃(연령 컬럼 없음)의 소제목행을 찾아 열 매핑 반환.

    소제목행 예: ['남자','여자',None,'300인 미만','300인 이상',None,'미적용','300인 미만','300인 이상']
    반환: (소제목행 index, {field: col}) 또는 (None, {}).
    """
    for ridx, row in enumerate(grid[:15]):
        cells = [(_bn(c) if c is not None else None) for c in row]
        if "남자" not in cells or "여자" not in cells:
            continue
        if not any(c and "300인" in c for c in cells if c):
            continue
        colmap = {"mort_m": cells.index("남자"), "mort_f": cells.index("여자")}
        mi = [j for j, c in enumerate(cells) if c and "300인미만" in c]
        ge = [j for j, c in enumerate(cells) if c and "300인이상" in c]
        napp = [j for j, c in enumerate(cells) if c == "미적용"]
        # 첫 번째 300인 쌍 = 퇴직률, 두 번째 쌍 = 승급률(미적용 이후)
        if mi:
            colmap["withdrawal_lt300"] = mi[0]
        if ge:
            colmap["withdrawal_ge300"] = ge[0]
        if napp:
            colmap["raise_none"] = napp[0]
        if len(mi) > 1:
            colmap["raise_lt300"] = mi[1]
        if len(ge) > 1:
            colmap["raise_ge300"] = ge[1]
        return ridx, colmap
    return None, {}


def _extract_base_year(grid) -> Optional[str]:
    """'2312 보험개발원 기초율' 등에서 기준연도(YYMM/YYYY) 추출."""
    for r in grid[:6]:
        for cell in r:
            if cell and "개발원" in str(cell):
                m = re.search(r"(\d{4})", str(cell))
                if m:
                    return m.group(1)
    return None


def _band_row(colmap, row) -> Dict:
    """밴드 열 매핑으로 한 데이터행을 리치 dict로 변환(값 있는 것만)."""
    def cell(field):
        c = colmap.get(field)
        return _num(row[c]) if c is not None and c < len(row) else None

    out = {"mort_m": cell("mort_m"), "mort_f": cell("mort_f")}
    for f in BAND_FIELDS:
        out[f] = cell(f)
    # 하위호환 단일필드: 미만 밴드를 기본으로
    out["withdrawal"] = out.get("withdrawal_lt300")
    out["raise_rate"] = (out.get("raise_lt300") if out.get("raise_lt300") is not None
                         else out.get("raise_none"))
    return out


def parse_base_rate_table(source: Union[str, Path, bytes], age_start: int = 15) -> Dict:
    """기초율표를 정규화 dict로 반환 — 헤더를 인식해 어떤 양식이든 맞춘다.

    반환: {
        'rows': [{'age':15,'withdrawal':...,'mort_m':...,'mort_f':...,'raise_rate':...,
                  (밴드형이면) 'withdrawal_lt300','withdrawal_ge300',
                  'raise_none','raise_lt300','raise_ge300'}, ...],
        'retirement_age': 60 | None,
        'avg_raise': 0.03 | None,
        'dev_format': True/False,      # 300인 밴드 포함 여부
        'base_year': '2312' | None,
    }
    표준(연령 컬럼) → 헤더 인식, 개발원(연령 없음) → 소제목 인식 후 age_start부터 채번.
    """
    from openpyxl import load_workbook

    if isinstance(source, (bytes, bytearray)):
        wb = load_workbook(io.BytesIO(source), data_only=True)
    else:
        wb = load_workbook(str(source), data_only=True)

    sheet = next((s for s in _SHEET_CANDIDATES if s in wb.sheetnames), wb.sheetnames[0])
    ws = wb[sheet]
    grid = list(ws.iter_rows(values_only=True))

    retirement_age = None
    avg_raise = None
    for r in grid[:8]:
        for j, cell in enumerate(r):
            if cell in ("정년", "정년연령") and j + 1 < len(r) and _num(r[j + 1]):
                retirement_age = int(_num(r[j + 1]))
            if cell in ("평균승급률", "평균임금상승률") and j + 1 < len(r):
                avg_raise = _num(r[j + 1])

    base_year = _extract_base_year(grid)

    # (A) 표준 헤더 우선 시도
    hdr_idx, colmap = _detect_header(grid)
    has_band = bool(colmap.keys() & set(BAND_FIELDS))
    rows: List[Dict] = []
    if colmap:
        def _cell(row, field):
            c = colmap.get(field)
            return _num(row[c]) if c is not None and c < len(row) else None
        for row in grid[hdr_idx + 1:]:
            age = _cell(row, "age")
            if age is None or not (10 <= age <= 110):
                if rows:
                    break
                continue
            rec = {
                "age": int(age),
                "withdrawal": _cell(row, "withdrawal"),
                "mort_m": _cell(row, "mort_m"),
                "mort_f": _cell(row, "mort_f"),
                "raise_rate": _cell(row, "raise_rate"),
            }
            if has_band:
                for f in BAND_FIELDS:
                    rec[f] = _cell(row, f)
                if rec["withdrawal"] is None:
                    rec["withdrawal"] = rec.get("withdrawal_lt300")
                if rec["raise_rate"] is None:
                    rec["raise_rate"] = (rec.get("raise_lt300")
                                         if rec.get("raise_lt300") is not None
                                         else rec.get("raise_none"))
            rows.append(rec)
        return {"rows": rows, "retirement_age": retirement_age, "avg_raise": avg_raise,
                "dev_format": has_band, "base_year": base_year}

    # (B) 개발원 소제목 레이아웃(연령 컬럼 없음)
    dev_idx, dev_map = _detect_dev_layout(grid)
    if dev_map:
        age = age_start
        for row in grid[dev_idx + 1:]:
            rec = _band_row(dev_map, row)
            # 데이터가 전부 비면 종료
            if all(v is None for v in rec.values()):
                if rows:
                    break
                continue
            rec["age"] = int(age)
            rows.append(rec)
            age += 1
        return {"rows": rows, "retirement_age": retirement_age, "avg_raise": avg_raise,
                "dev_format": True, "base_year": base_year}

    # (C) 위치 기반 폴백(A~E: 연령·퇴직률·사망률남·사망률여·승급률)
    for r in grid:
        age = _num(r[0]) if r else None
        if age is None or not (10 <= age <= 110):
            if rows:
                break
            continue
        rows.append({
            "age": int(age),
            "withdrawal": _num(r[1]) if len(r) > 1 else None,
            "mort_m": _num(r[2]) if len(r) > 2 else None,
            "mort_f": _num(r[3]) if len(r) > 3 else None,
            "raise_rate": _num(r[4]) if len(r) > 4 else None,
        })
    return {"rows": rows, "retirement_age": retirement_age, "avg_raise": avg_raise,
            "dev_format": False, "base_year": base_year}


def band_withdrawal(row: Dict, size_band: str) -> Optional[float]:
    """행에서 사업장 규모 밴드('lt300'/'ge300')에 맞는 퇴직률을 고른다."""
    if size_band == "ge300" and row.get("withdrawal_ge300") is not None:
        return row["withdrawal_ge300"]
    if size_band == "lt300" and row.get("withdrawal_lt300") is not None:
        return row["withdrawal_lt300"]
    # 밴드 컬럼이 없으면 단일 퇴직률
    if row.get("withdrawal") is not None:
        return row["withdrawal"]
    return row.get("withdrawal_lt300") if size_band != "ge300" else row.get("withdrawal_ge300")


def band_raise(row: Dict, size_band: str) -> Optional[float]:
    """행에서 사업장 규모 밴드에 맞는 승급률을 고른다(밴드 없으면 단일/미적용)."""
    if size_band == "ge300" and row.get("raise_ge300") is not None:
        return row["raise_ge300"]
    if size_band == "lt300" and row.get("raise_lt300") is not None:
        return row["raise_lt300"]
    if row.get("raise_rate") is not None:
        return row["raise_rate"]
    return row.get("raise_none")


def build_base_rate_template() -> bytes:
    """기초율 입력 표준 양식(작성요령 + 개발원식 300인 밴드 헤더 + 예시행) xlsx 바이트 생성."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "작성요령"
    ws.column_dimensions["A"].width = 96
    ws["A1"] = "[기초율 입력 양식] 작성요령 — 보험개발원 기초율(300인 미만/이상 구분)"
    ws["A1"].font = Font(name="맑은 고딕", size=13, bold=True, color="1F4E79")
    guide = [
        "· '기초율표' 시트에 연령별로 값을 입력하세요. 컬럼 순서·이름이 조금 달라도 헤더로 자동 인식합니다.",
        "· 연령: 15~정년 등 표에 넣을 연령. 한 행에 한 연령.",
        "· 사망률(남)/(여): 성별·연령별 사망률(소수, 예: 0.003% → 0.00003).",
        "· 퇴직률은 사업장 규모별로 다릅니다 → 300인 미만 / 300인 이상 두 열에 각각 입력.",
        "· 승급률(임금상승률)도 미적용 / 300인 미만 / 300인 이상으로 나누어 입력(소수, 예: 5% → 0.05).",
        "· 상단 셀에 정년(예: 60)을 적으면 세트 기본값으로 함께 저장됩니다.",
        "· 개발원 원본(연령 열이 없고 남자/여자·300인 미만/이상 소제목만 있는 파일)도 그대로 업로드하면 "
        "  시작연령(기본 15세)부터 자동 인식합니다.",
        "· 개발원 기초율은 3년 주기로 변경됩니다. 변경 시 새 세트(버전)로 업로드하세요.",
    ]
    for i, line in enumerate(guide, start=3):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(name="맑은 고딕", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    ws2 = wb.create_sheet("기초율표")
    ws2["A1"] = "정년"; ws2["B1"] = 60
    headers = ["연령", "사망률(남)", "사망률(여)",
               "퇴직률(300인미만)", "퇴직률(300인이상)",
               "승급률(미적용)", "승급률(300인미만)", "승급률(300인이상)"]
    fill = PatternFill("solid", fgColor="DCE6F1")
    hf = Font(name="맑은 고딕", size=10, bold=True)
    for j, h in enumerate(headers, start=1):
        cell = ws2.cell(row=3, column=j, value=h)
        cell.font = hf
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.column_dimensions[chr(64 + j)].width = 16
    # 예시 몇 행 (연령, 사망남, 사망여, 퇴직<300, 퇴직>=300, 승급미적용, 승급<300, 승급>=300)
    example = [
        (15, 0.00003, 0.00002, 0.39444, 0.17236, 0.0, 0.04474, 0.06743),
        (30, 0.00020, 0.00010, 0.15000, 0.10000, 0.0, 0.05000, 0.06000),
        (50, 0.00088, 0.00053, 0.06000, 0.05000, 0.0, 0.02000, 0.02500),
    ]
    for i, row in enumerate(example, start=4):
        for j, v in enumerate(row, start=1):
            ws2.cell(row=i, column=j, value=v)
    ws2.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
