"""통합 테스트: 샘플 500명 명부로 전체 파이프라인이 끝까지 돌아가는지 확인.

명부 생성(한글 컬럼) → 컬럼 매핑 로딩 → 검증 → 계산 → 산출물(xlsx/json) 저장.
"""

import json
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))
from generate_sample_census import generate  # noqa: E402

from dbo.census import load_census, validate_census  # noqa: E402
from dbo.config import Config  # noqa: E402
from dbo.decrement import DecrementTables  # noqa: E402
from dbo.engine import calculate_census  # noqa: E402
from dbo.outputs import write_outputs  # noqa: E402

CONFIG_PATH = "config/assumptions_sample.yaml"
COLMAP_PATH = "config/column_map_sample.yaml"


@pytest.fixture
def sample_census_file(tmp_path):
    df = generate(n=500, seed=7)
    path = tmp_path / "sample_census.xlsx"
    df.to_excel(path, index=False)
    return path


def test_full_pipeline_runs_end_to_end(tmp_path, sample_census_file):
    config = Config.from_yaml(CONFIG_PATH)
    tables = DecrementTables.from_config(config, base_dir="config")

    # 한글 컬럼 → 표준 스키마 매핑 로딩
    records, report, df = load_census(sample_census_file, column_map=COLMAP_PATH)
    assert len(records) > 0
    assert "emp_id" in df.columns          # 매핑되어 표준 컬럼명 존재
    assert set(df["gender"].unique()) <= {"M", "F"}   # 값 매핑(남/여→M/F)

    validate_census(records, config.valuation_date, report)
    # 더미 데이터엔 치명 오류가 없어야 계산까지 진행 (경고는 허용)
    assert not report.has_errors, [i.message for i in report.errors][:5]

    result = calculate_census(records, config, tables, with_detail=False)
    assert result.total_dbo > 0
    assert len(result.results) + len(result.excluded_emp_ids) == len(records)
    # 제도구분 3은 제외 목록으로
    assert all(r.plan_type != 3 for r in result.results)

    out_dir = tmp_path / "results"
    paths = write_outputs(out_dir, records, result, config, tables,
                          census_path=sample_census_file, report=report)

    # 산출물 파일 생성 확인
    assert paths["xlsx"].exists()
    assert paths["run_log"].exists()

    # 엑셀 시트 구성 확인
    wb = openpyxl.load_workbook(paths["xlsx"])
    for sheet in ["개인별산출표", "요약", "민감도분석", "만기분석"]:
        assert sheet in wb.sheetnames

    # 실행 로그 JSON 내용 확인
    log = json.loads(paths["run_log"].read_text(encoding="utf-8"))
    assert log["results"]["total_dbo"] == result.total_dbo
    assert len(log["input"]["census_sha256"]) == 64      # SHA-256 hex
    assert "config_snapshot" in log


def test_report_sheets_and_maturity_reconciles(tmp_path, sample_census_file):
    """전문 계리평가보고서(사업보고서 서식)의 8개 시트가 생성되고,
    만기구성표 총계가 재무상태표상 확정급여채무(total_dbo)와 일치하는지 확인."""
    from dbo.engine import expected_cashflows

    config = Config.from_yaml(CONFIG_PATH)
    tables = DecrementTables.from_config(config, base_dir="config")
    records, report, _ = load_census(sample_census_file, column_map=COLMAP_PATH)
    validate_census(records, config.valuation_date, report)
    result = calculate_census(records, config, tables, with_detail=False)

    out_dir = tmp_path / "results"
    di = {"dbo_begin": 1.8e9, "plan_assets": 1.2e9, "plan_assets_begin": 1.1e9,
          "contributions": 1e8, "interest_income": 4.4e7, "net_interest": 2.8e7}
    paths = write_outputs(out_dir, records, result, config, tables,
                          census_path=sample_census_file, report=report,
                          company="샘플전자(주)", disclosure_inputs=di)

    assert paths["report"].exists()
    wb = openpyxl.load_workbook(paths["report"])
    for sheet in ["표지", "목차", "의견서", "Ⅰ.개요·방법론", "Ⅱ.주석공시사항",
                  "Ⅳ.기타세부내역", "참고", "개인별명세"]:
        assert sheet in wb.sheetnames

    # 만기별 PUC 배분 현재가치 합 == 제도1(DB정상평가) DBO (연도별 반올림 오차 허용)
    cf = expected_cashflows(records, config, tables)
    puc_dbo = result.subtotal_by_plan.get(1, {}).get("DBO", 0.0)
    assert cf["현재가치"].sum() == pytest.approx(puc_dbo, rel=1e-6, abs=len(records))

    # 만기구성표 총계(간편법 조정행 포함) == total_dbo
    simple_dbo = result.subtotal_by_plan.get(2, {}).get("DBO", 0.0)
    assert cf["현재가치"].sum() + simple_dbo == pytest.approx(result.total_dbo, abs=len(records))


def test_cli_run_smoke(tmp_path, sample_census_file):
    from dbo.cli import main

    out_dir = tmp_path / "cli_results"
    rc = main([
        "run",
        "--census", str(sample_census_file),
        "--config", CONFIG_PATH,
        "--map", COLMAP_PATH,
        "--out", str(out_dir),
        "--debug-emp", "10000",
    ])
    assert rc == 0
    assert (out_dir / "dbo_results.xlsx").exists()
    assert (out_dir / "run_log.json").exists()
