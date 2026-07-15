"""보조 양식 3종(사외적립·기타장기·명부확인요약표) — 사용자 원본 양식 기반 테스트."""

import io

import openpyxl

from dbo import aux_forms as AF


def _fill(bytes_in, edits):
    """양식 바이트를 열어 (row,col)=value 로 채운 뒤 다시 바이트로."""
    wb = openpyxl.load_workbook(io.BytesIO(bytes_in))
    ws = wb[wb.sheetnames[0]]
    for (r, c), v in edits.items():
        ws.cell(row=r, column=c, value=v)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def test_funding_template_is_user_sheet_and_roundtrip():
    b = AF.build_funding_template("2025-12-31")
    wb = openpyxl.load_workbook(io.BytesIO(b))
    assert wb.sheetnames == ["사외적립자산"]        # 사용자 원본 시트만
    # 입력은 F열: 기초잔액(F5)·입금액(F6)·퇴직(F7)
    filled = _fill(b, {(5, 6): 1_000_000_000, (6, 6): 200_000_000, (7, 6): 50_000_000})
    out = AF.parse_funding_upload(filled)
    assert out["기초잔액"] == 1_000_000_000
    assert out["입금액"] == 200_000_000
    assert out["지급_퇴직"] == 50_000_000


def test_other_lt_template_is_user_sheet():
    b = AF.build_other_lt_template()
    wb = openpyxl.load_workbook(io.BytesIO(b))
    assert wb.sheetnames == ["기타 장기"]           # 사용자 원본 시트만


class _Rec:
    def __init__(self, emp_class, accrual, interim_date=None, interim_amt=None):
        self.emp_class = emp_class
        self.current_year_accrual = accrual
        self.interim_settlement_date = interim_date
        self.interim_settlement_amount = interim_amt


def test_compute_census_summary():
    import datetime as dt
    recs = [
        _Rec("EXECUTIVE", 100_000_000),
        _Rec("REGULAR", 50_000_000, dt.date(2023, 1, 1), 10_000_000),
        _Rec("REGULAR", 30_000_000),
        _Rec("CONTRACT", 20_000_000),
    ]
    s = AF.compute_census_summary(recs)
    assert s["재직_임원"] == 1
    assert s["재직_직원"] == 2
    assert s["재직_계약직"] == 1
    assert s["재직_합계"] == 4
    assert s["추계액_임원"] == 100_000_000
    assert s["추계액_직원"] == 80_000_000
    assert s["추계액_합계"] == 200_000_000
    assert s["중간정산자수"] == 1
    assert s["중간정산금액"] == 10_000_000


def test_census_summary_template_autofill_g_column():
    s = {"재직_임원": 1, "재직_직원": 2, "재직_계약직": 1, "재직_합계": 4,
         "추계액_임원": 1e8, "추계액_직원": 8e7, "추계액_계약직": 2e7, "추계액_합계": 2e8,
         "중간정산자수": 1, "중간정산금액": 1e7}
    b = AF.build_census_summary_template(s, "2025-12-31")
    wb = openpyxl.load_workbook(io.BytesIO(b))
    assert wb.sheetnames == ["명부확인용요약표"]
    ws = wb["명부확인용요약표"]
    # G열 자동산출: 재직합계(G9)=4, 추계액합계(G19)=2e8, 중간정산자수(G14)=1
    assert ws["G9"].value == 4
    assert ws["G19"].value == 200000000
    assert ws["G14"].value == 1
    # 파싱 round-trip: F열(회사입력) 한 칸 채워 파싱
    filled = _fill(b, {(6, 6): 999})   # F6
    out = AF.parse_census_summary_upload(filled)
    assert out["rows"], "요약표 행이 파싱되어야 함"


def test_discount_upload_template_roundtrip():
    b = AF.build_discount_upload_template()
    wb = openpyxl.load_workbook(io.BytesIO(b))
    ws = wb.active
    ws["B1"] = "202512"
    c0 = 1 + 5 * 3            # A0 등급 시작열
    ws.cell(row=5, column=c0 + 1, value="3.16%")
    ws.cell(row=6, column=c0 + 1, value="3.30%")
    ws.cell(row=7, column=c0 + 1, value="3.49%")
    ws.cell(row=5, column=7 + 1, value=0.030)   # AA0 만기1 = 0.03
    buf = io.BytesIO(); wb.save(buf)
    out = AF.parse_discount_upload(buf.getvalue())
    assert out["기준일"] == "202512"
    assert "A0" in out["curves"] and "AA0" in out["curves"]
    a0 = out["curves"]["A0"]
    assert a0[0] == {"maturity": 1, "rate": 0.0316}
    assert len(a0) == 3


def test_parse_simple_curve():
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["만기", "할인율"])
    for m, r in [(1, "3.1%"), (3, "3.4%"), (5, "3.8%")]:
        ws.append([m, r])
    buf = io.BytesIO(); wb.save(buf)
    pts = AF.parse_simple_curve(buf.getvalue())
    assert pts == [{"maturity": 1, "rate": 0.031}, {"maturity": 3, "rate": 0.034},
                   {"maturity": 5, "rate": 0.038}]
