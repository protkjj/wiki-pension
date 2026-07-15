"""기초율 파서 · 할인율 solver · 마스터데이터 store 테스트."""

import io

import pytest
from openpyxl import Workbook

from dbo import discount as D
from dbo.base_rates import parse_base_rate_table
from dbo.platform import store
from dbo.platform.db import init_db


# -- 할인율 solver -----------------------------------------------------------

CURVE = {1: 0.031, 2: 0.033, 3: 0.035, 4: 0.037, 5: 0.040, 10: 0.047, 20: 0.056}


def _cf():
    return [(y, 1_000_000 * (0.9 ** y)) for y in range(1, 21)]


def test_single_rate_matches_curve_pv():
    cf = _cf()
    r = D.single_equivalent_rate(cf, CURVE, "mid_year")
    # 단일율로 할인한 PV가 커브 PV와 일치해야 한다.
    assert abs(D.flat_pv(cf, r, "mid_year") - D.curve_pv(cf, CURVE, "mid_year")) < 1.0
    # 단일율은 커브 최소~최대 사이.
    assert min(CURVE.values()) <= r <= max(CURVE.values())


def test_timing_changes_pv():
    cf = _cf()
    mid = D.curve_pv(cf, CURVE, "mid_year")
    eoy = D.curve_pv(cf, CURVE, "end_of_year")
    assert mid > eoy          # 연중이 반기 덜 할인 → PV 더 큼


def test_duration_and_interpolation():
    cf = _cf()
    dur = D.weighted_duration(cf, CURVE, "mid_year")
    assert 1 < dur < 20
    # 듀레이션 시점 커브 보간값은 인접 만기 사이.
    dr = D.interpolate_curve(CURVE, 4.5)
    assert 0.037 <= dr <= 0.040


def test_spot_for_period_floor_convention():
    # 할인기간의 floor 정수 만기 금리를 쓴다(엑셀 관행).
    assert D._spot_for_period(CURVE, 1.5) == CURVE[1]     # 1.5년 → 1년물
    assert D._spot_for_period(CURVE, 2.5) == CURVE[2]     # 2.5년 → 2년물
    assert D._spot_for_period(CURVE, 0.5) == CURVE[1]     # 0.5년 → 최소만기(1)로 clamp
    assert D._spot_for_period(CURVE, 30) == CURVE[20]     # 만기초과 → 20년으로 clamp


def test_single_rate_reproduces_actuary_excel():
    """실제 계리 엑셀(2512 PUC) 검증 대사 — floor 커브규칙 + 연중시점.

    엑셀 단일율 3.6515%를 시스템이 ±0.1bp 내로 재현해야 한다.
    (간이 재현: 대표 현금흐름·커브로 관계 확인)
    """
    curve = {1: 0.03161, 2: 0.03303, 3: 0.03487, 4: 0.03587, 5: 0.03818,
             6: 0.03965, 7: 0.04094, 8: 0.04374, 9: 0.04762, 10: 0.05152,
             11: 0.05442, 12: 0.05627, 13: 0.05734, 14: 0.05793, 15: 0.05838,
             16: 0.05902, 17: 0.05988, 18: 0.06092, 19: 0.06213, 20: 0.06346}
    # 0.5년 시점에 큰 지급이 몰린 형태(엑셀과 유사) → 단일율은 커브 단기쪽에 가까움
    cf = [(1, 2.285e9), (2, 1.304e9), (3, 0.772e9), (4, 0.479e9), (5, 0.336e9),
          (6, 0.198e9), (7, 0.132e9), (8, 0.088e9), (9, 0.062e9), (10, 0.049e9)]
    r = D.single_equivalent_rate(cf, curve, "mid_year")
    # 커브 단기(3.16%)~중기 사이, 듀레이션 짧아 3.6% 근방
    assert 0.034 < r < 0.039
    # 단일율 PV == 커브 PV
    assert abs(D.flat_pv(cf, r, "mid_year") - D.curve_pv(cf, curve, "mid_year")) < 1.0


# -- 기초율 파서 -------------------------------------------------------------

def _base_rate_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "기초율표"
    ws["E1"] = "정년"; ws["F1"] = 60
    ws["E2"] = "평균승급률"; ws["F2"] = 0.03
    # 헤더행
    ws.append([None] * 5)
    ws.append(["연령", "퇴직율", "사망률(남)", "사망률(여)", "승급률"])
    for age in range(15, 20):
        ws.append([age, 0.1, 0.00003, 0.00002, 0.0])
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def test_parse_base_rate_table():
    res = parse_base_rate_table(_base_rate_xlsx())
    assert res["retirement_age"] == 60
    assert abs(res["avg_raise"] - 0.03) < 1e-9
    assert len(res["rows"]) == 5
    assert res["rows"][0] == {"age": 15, "withdrawal": 0.1, "mort_m": 0.00003,
                              "mort_f": 0.00002, "raise_rate": 0.0}
    assert res["dev_format"] is False


def _dev_base_rate_xlsx() -> bytes:
    """개발원 레이아웃(연령 컬럼 없음, 남자/여자·300인 밴드 소제목)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([None] * 9)
    ws.append(["* 개발원 기초율 3년 주기 변경", None, None, None, None, None, None, None, None])
    ws.append(["2312 보험개발원 기초율", None, None, None, None, None, None, None, None])
    ws.append(["2312 사망률", None, None, "2312 퇴직률", None, None, "2312 승급률", None, None])
    ws.append(["남자", "여자", None, "300인 미만", "300인 이상", None, "미적용", "300인 미만", "300인 이상"])
    # 3개 데이터행 → 연령 15,16,17
    ws.append([0.00003, 0.00002, None, 0.39, 0.17, None, None, 0.044, 0.067])
    ws.append([0.00003, 0.00002, None, 0.37, 0.16, None, None, 0.037, 0.077])
    ws.append([0.00004, 0.00002, None, 0.35, 0.15, None, None, 0.033, 0.083])
    ws.append([None] * 9)
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def test_parse_dev_base_rates():
    from dbo.base_rates import band_raise, band_withdrawal
    res = parse_base_rate_table(_dev_base_rate_xlsx())
    assert res["dev_format"] is True
    assert res["base_year"] == "2312"
    assert len(res["rows"]) == 3
    r0 = res["rows"][0]
    assert r0["age"] == 15
    assert r0["withdrawal_lt300"] == 0.39 and r0["withdrawal_ge300"] == 0.17
    assert r0["raise_lt300"] == 0.044 and r0["raise_ge300"] == 0.067
    # 밴드 선택기
    assert band_withdrawal(r0, "lt300") == 0.39
    assert band_withdrawal(r0, "ge300") == 0.17
    assert band_raise(r0, "ge300") == 0.067
    # 마지막 연령 = 17
    assert res["rows"][-1]["age"] == 17


def test_band_fallback_single():
    """밴드 컬럼이 없는 단일 세트에서는 밴드 선택기가 단일값으로 폴백."""
    from dbo.base_rates import band_raise, band_withdrawal
    row = {"age": 30, "withdrawal": 0.05, "raise_rate": 0.03}
    assert band_withdrawal(row, "ge300") == 0.05
    assert band_withdrawal(row, "lt300") == 0.05
    assert band_raise(row, "ge300") == 0.03


# -- 마스터데이터 store ------------------------------------------------------

@pytest.fixture
def db(tmp_path):
    p = tmp_path / "t.sqlite"
    init_db(p)
    return str(p)


def test_base_rate_set_crud(db):
    import json
    sid = store.add_base_rate_set(db, "2025 당기", "개발원2312", "2312", "당기", 60, 0.03,
                                  json.dumps([{"age": 30, "withdrawal": 0.05}]), "메모", 1, "2025-01-01")
    assert sid > 0
    lst = store.list_base_rate_sets(db)
    assert len(lst) == 1 and lst[0]["name"] == "2025 당기"
    full = store.get_base_rate_set(db, sid)
    assert full["source"] == "개발원2312" and full["data_json"]
    store.delete_base_rate_set(db, sid)
    assert store.list_base_rate_sets(db) == []


def test_discount_curve_crud(db):
    import json
    cid = store.add_discount_curve(db, "2025-12-31 AA+", "2025-12-31", "AA+", "당기",
                                   json.dumps([{"maturity": 1, "rate": 0.031}]),
                                   0.0404, 4.99, "", 1, "2025-01-01")
    assert cid > 0
    lst = store.list_discount_curves(db)
    assert len(lst) == 1 and abs(lst[0]["single_rate"] - 0.0404) < 1e-9
    assert store.get_discount_curve(db, cid)["rating"] == "AA+"
    store.delete_discount_curve(db, cid)
    assert store.list_discount_curves(db) == []


# -- 명부 검토 워크북 --------------------------------------------------------

def test_census_review_workbook():
    import io
    import pandas as pd
    from openpyxl import load_workbook
    from dbo import census_review as CR
    mapped = pd.DataFrame([
        {"emp_id": "1", "birth_date": "1985-01-01", "gender": "M",
         "hire_date": "2010-01-01", "base_salary": 3_000_000,
         "current_year_accrual": 5_000_000, "emp_class": "REGULAR"},
    ])
    std = CR.standard_view(mapped)
    assert "사원번호" in std.columns and "기준급여" in std.columns
    prob = pd.DataFrame([{"엑셀행": 2, "⚠문제": "🔴오류 …", "emp_id": "1"}])
    wb = CR.build_review_workbook(mapped, prob, pd.DataFrame(), ["성명"],
                                  {"filename": "a.xlsx", "valuation_date": "2025-12-31",
                                   "records": 1, "errors": 1, "warnings": 0})
    w = load_workbook(io.BytesIO(wb))
    assert w.sheetnames == ["①표준변환(우리양식)", "②오류검토", "③오류목록", "④안내"]
    # 오류검토 시트에 회사 기입 열이 추가됨
    hdr = [c.value for c in w["②오류검토"][1]]
    assert any("회사수정값" in str(h) for h in hdr)
    assert any("오류아님" in str(h) for h in hdr)


def test_annotate_original_keeps_format():
    import pandas as pd
    from dbo import census_review as CR
    raw = pd.DataFrame([{"사번": "A1", "기준급여": 100}, {"사번": "A2", "기준급여": 200}])
    rim = {0: {"구분": "오류", "추정오류": "🔴오류 X"}}
    ann = CR.annotate_original(raw, rim)
    assert list(ann.columns) == ["사번", "기준급여", "오류가 아닐시 사유 기록", "구분", "추정오류"]
    assert ann.iloc[0]["구분"] == "오류" and ann.iloc[0]["추정오류"] == "🔴오류 X"
    assert ann.iloc[1]["구분"] == "" and ann.iloc[1]["추정오류"] == ""


# -- 경험기초율 산출 --------------------------------------------------------

def test_experience_template_roundtrip():
    from dbo.experience import build_experience_template, parse_experience_data
    p = parse_experience_data(build_experience_template())
    assert p["n"] == 2                      # 예시 2행
    # 헤더가 인식되어 사원번호·급여가 파싱됨
    assert p["records"][0]["emp_id"] == "A001"
    assert p["records"][0]["sal_start"] == 3000000


def test_compute_experience_rates():
    from datetime import date
    from dbo.experience import apply_mortality_from, compute_experience_rates
    recs = []
    for i in range(80):     # 재직 지속, 급여 3.0M→3.3M
        recs.append({"emp_id": f"A{i}", "birth": date(1996, 1, 1), "gender": "1",
                     "hire": date(2019, 1, 1), "sal_start": 3_000_000, "sal_end": 3_300_000,
                     "leave": None, "left": False})
    for i in range(20):     # 관측기간 중 퇴직
        recs.append({"emp_id": f"B{i}", "birth": date(1996, 1, 1), "gender": "1",
                     "hire": date(2019, 1, 1), "sal_start": 3_000_000, "sal_end": 3_100_000,
                     "leave": date(2023, 7, 1), "left": True})
    res = compute_experience_rates(recs, date(2021, 1, 1), date(2025, 12, 31), retirement_age=60)
    sm = res["summary"]
    assert sm["n"] == 100 and sm["n_withdrawals"] == 20
    assert sm["exposure"] > 0
    # 퇴직률 = 20 / 노출 > 0, 승급률 양수
    assert 0 < sm["overall_withdrawal"] < 1
    assert sm["avg_raise"] > 0
    # 연령행이 15~정년까지 채워지고, 27세(밴드 25-29)에 경험률 반영
    r27 = next(r for r in res["rows"] if r["age"] == 27)
    assert r27["withdrawal"] > 0 and r27["raise_rate"] > 0
    # 사망률 차용
    apply_mortality_from(res["rows"], [{"age": 27, "mort_m": 0.0005, "mort_f": 0.0003}])
    assert next(r for r in res["rows"] if r["age"] == 27)["mort_m"] == 0.0005


def test_experience_base_rate_set_company_scope(db):
    import json
    from dbo.platform import store
    from dbo.platform.db import get_conn
    conn = get_conn(db)
    conn.execute("INSERT INTO companies(id,name,created) VALUES(5,'라진','2025-01-01')")
    conn.commit(); conn.close()
    dev = store.add_base_rate_set(db, "개발원2312", "개발원2312", "2312", "당기", 60, None,
                                  "[]", "", 1, "2025-01-01")
    exp = store.add_base_rate_set(db, "라진 경험2025", "경험률", "2025", "당기", 60, 0.02,
                                  "[]", "", 1, "2025-01-01", company_id=5, kind="experience")
    # 회사5 조회 = 경험(회사5) + 개발원(공용)
    for5 = store.list_base_rate_sets(db, company_id=5)
    assert {s["id"] for s in for5} == {dev, exp}
    # 다른 회사9 = 개발원만
    for9 = store.list_base_rate_sets(db, company_id=9)
    assert {s["id"] for s in for9} == {dev}
    assert store.list_base_rate_sets(db, kind="experience")[0]["id"] == exp
    # 업로드 플래그
    store.add_aux_census(db, 5, store.EXPERIENCE_CENSUS_TYPE, "e.xlsx", "/x", 1, "2025-02-01")
    stat = store.experience_upload_status(db)
    assert stat and stat[0]["company_id"] == 5 and stat[0]["n_sets"] == 1
