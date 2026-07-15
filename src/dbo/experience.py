"""경험기초율 산출 — 기업 제공 데이터로 경험 퇴직률·승급률을 추정한다.

계리업의 핵심인 경험기초율(경험률)은 회사 고유의 과거 이력에서 연령대별
퇴직률과 승급률(임금상승률)을 직접 추정한 값이다. 개발원 표준율과 달리
회사의 실제 탈퇴·급여상승 경험을 반영한다.

입력(기업이 올리는 '경험기초율 산출데이터'):
  관측기간(관측시작일~관측종료일) 동안의 종업원별 1행 —
  사원번호 · 생년월일 · 성별 · 입사일자 · 관측시작급여 · 관측종료급여 · 상태(재직/퇴직) · 퇴직일자

산출(연령대 5세 밴드):
  · 경험퇴직률 q = Σ퇴직건 / Σ노출(인·년)          (중심탈퇴율 근사)
  · 경험승급률 g = 재직자의 연환산 급여상승률 평균   ( (기말/기초)^(1/연) − 1 )
  밴드값을 연령(15~정년)으로 펼쳐 기초율표 행으로 만든다(사망률은 개발원에서 차용).
"""

from __future__ import annotations

import io
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Union

from .census import _emp_id_to_str, _norm_header

# 경험데이터 컬럼 별칭(헤더 자동 인식)
_COLS = {
    "emp_id": ["사원번호", "사번", "직원번호", "사원코드", "empid", "id"],
    "birth": ["생년월일", "생일", "출생일", "birth"],
    "gender": ["성별", "gender", "sex"],
    "hire": ["입사일자", "입사일", "hire"],
    "sal_start": ["관측시작급여", "시작급여", "기초급여", "전기급여", "기초월급여"],
    "sal_end": ["관측종료급여", "종료급여", "기말급여", "당기급여", "기말월급여", "현재급여"],
    "status": ["상태", "재직여부", "재직상태", "status"],
    "leave": ["퇴직일자", "퇴직일", "탈퇴일", "leave"],
}
_OBS_LABELS = {"start": ["관측시작일", "관측개시일", "시작일"],
               "end": ["관측종료일", "관측말일", "종료일", "기준일"]}


def _to_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _emp_id_to_str(v)
    if not s:
        return None
    s = s.replace("-", "").replace("/", "").replace(".", "").strip()
    if len(s) == 8 and s.isdigit():
        try:
            return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        except ValueError:
            return None
    return None


def _to_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return None


def _is_left(status_val, leave_date) -> bool:
    """상태값 또는 퇴직일 존재로 '퇴직' 여부 판정 (재직=1/재직, 퇴직=2/퇴직/Y)."""
    s = _emp_id_to_str(status_val)
    if s is not None:
        s = s.strip().lower()
        if s in ("2", "퇴직", "탈퇴", "y", "leave", "left", "resigned"):
            return True
        if s in ("1", "재직", "n", "active", "stay"):
            return False
    return leave_date is not None


def parse_experience_data(source: Union[str, Path, bytes]) -> Dict:
    """경험데이터 xlsx/csv를 파싱 — 헤더 자동 인식. 관측기간 셀도 함께 읽는다.

    반환: {'records': [{emp_id,birth,gender,hire,sal_start,sal_end,left,leave}],
           'obs_start': date|None, 'obs_end': date|None, 'n': int}
    """
    from openpyxl import load_workbook

    if isinstance(source, (bytes, bytearray)):
        wb = load_workbook(io.BytesIO(source), data_only=True)
    else:
        wb = load_workbook(str(source), data_only=True)
    # 데이터 시트 우선(작성요령 제외)
    sheet = next((s for s in wb.sheetnames if "요령" not in s), wb.sheetnames[0])
    ws = wb[sheet]
    grid = list(ws.iter_rows(values_only=True))

    # 관측기간 셀(라벨-값) 스캔
    obs_start = obs_end = None
    for r in grid[:12]:
        for j, cell in enumerate(r or []):
            if cell is None:
                continue
            n = _norm_header(cell)
            if any(_norm_header(a) == n for a in _OBS_LABELS["start"]) and j + 1 < len(r):
                obs_start = obs_start or _to_date(r[j + 1])
            if any(_norm_header(a) == n for a in _OBS_LABELS["end"]) and j + 1 < len(r):
                obs_end = obs_end or _to_date(r[j + 1])

    # 헤더행(사원번호 포함) 찾기
    alias_norm = {f: [_norm_header(a) for a in al] for f, al in _COLS.items()}
    hdr_idx, colmap = None, {}
    for ridx, row in enumerate(grid[:15]):
        norm = {}
        for cidx, cell in enumerate(row or []):
            if cell is not None:
                norm.setdefault(_norm_header(cell), cidx)
        cm = {}
        for field, na in alias_norm.items():
            for a in na:
                if a in norm:
                    cm[field] = norm[a]
                    break
        if "emp_id" in cm and ("birth" in cm or "sal_start" in cm):
            hdr_idx, colmap = ridx, cm
            break
    records: List[Dict] = []
    if hdr_idx is not None:
        for row in grid[hdr_idx + 1:]:
            def g(field):
                c = colmap.get(field)
                return row[c] if c is not None and c < len(row) else None
            emp = _emp_id_to_str(g("emp_id"))
            if not emp:
                continue
            leave = _to_date(g("leave"))
            records.append({
                "emp_id": emp,
                "birth": _to_date(g("birth")),
                "gender": _emp_id_to_str(g("gender")),
                "hire": _to_date(g("hire")),
                "sal_start": _to_float(g("sal_start")),
                "sal_end": _to_float(g("sal_end")),
                "leave": leave,
                "left": _is_left(g("status"), leave),
            })
    return {"records": records, "obs_start": obs_start, "obs_end": obs_end, "n": len(records)}


def _band_of(age: int, width: int = 5, lo: int = 15, hi: int = 60) -> int:
    """연령 → 밴드 하한(예 5세폭: 27→25). 경계 밖은 최소/최대 밴드로 clamp.

    밴드 집합은 range(lo, hi, width)라 상한 밴드는 hi를 포함하지 않는다. 정년(hi)에
    정확히 걸리는 연령은 마지막 밴드로 접어 넣는다(정년 도달자는 이 밴드에 귀속) —
    그렇지 않으면 존재하지 않는 밴드 키로 KeyError가 난다.
    """
    age = max(lo, min(hi, age))
    band = ((age - lo) // width) * width + lo
    top_band = lo + max(0, (hi - 1 - lo) // width) * width   # range(lo,hi,width)의 마지막 밴드
    return min(band, top_band)


# 퇴직자명부(직전 3~5년) 컬럼 별칭 — 경험퇴직률 산출용
_RET_COLS = {
    "emp_id": ["사원번호", "사번", "직원번호", "사원코드"],
    "birth": ["생년월일", "생일", "출생일"],
    "gender": ["성별", "gender", "sex"],
    "hire": ["입사일자", "입사일", "hire"],
    "leave": ["퇴직일_DC전환일", "퇴직일DC전환일", "퇴직일또는DC전환일", "퇴직일", "탈퇴일", "퇴직일자"],
    "amount": ["퇴직금_DC전환금", "퇴직금DC전환금", "퇴직금또는DC전환금", "퇴직금", "DC전환금", "지급액"],
    "emp_class": ["종업원구분", "종업원구분", "구분", "직군", "직종"],
    "reason": ["사유", "reason"],
}


def parse_retiree_census(source: Union[str, Path, bytes]) -> List[Dict]:
    """퇴직자명부 xlsx/csv 파싱 → [{emp_id,birth,gender,hire,leave,amount,emp_class,reason}].

    헤더 자동 인식(괄호주석·줄바꿈 무시). 경험퇴직률 산출의 분자(퇴직 건수) 데이터.
    """
    from openpyxl import load_workbook

    if isinstance(source, (bytes, bytearray)):
        wb = load_workbook(io.BytesIO(source), data_only=True)
    else:
        wb = load_workbook(str(source), data_only=True)
    sheet = next((s for s in wb.sheetnames if "요령" not in s), wb.sheetnames[0])
    grid = list(wb[sheet].iter_rows(values_only=True))

    alias = {f: [_norm_header(a) for a in al] for f, al in _RET_COLS.items()}
    hdr_idx, colmap = None, {}
    for ridx, row in enumerate(grid[:15]):
        cm = {}
        norm = {}
        for cidx, cell in enumerate(row or []):
            if cell is not None:
                norm.setdefault(_norm_header(cell), cidx)
        for field, na in alias.items():
            for a in na:
                if a in norm:
                    cm[field] = norm[a]
                    break
        if "emp_id" in cm and "leave" in cm:
            hdr_idx, colmap = ridx, cm
            break
    if hdr_idx is None:
        return []

    out = []
    for row in grid[hdr_idx + 1:]:
        if not row:
            continue

        def g(f):
            j = colmap.get(f)
            return row[j] if (j is not None and j < len(row)) else None

        emp = _emp_id_to_str(g("emp_id"))
        birth = _to_date(g("birth"))
        leave = _to_date(g("leave"))
        if emp is None and birth is None and leave is None:
            continue
        out.append({
            "emp_id": emp, "birth": birth, "gender": _emp_id_to_str(g("gender")),
            "hire": _to_date(g("hire")), "leave": leave,
            "amount": _to_float(g("amount")),
            "emp_class": _emp_id_to_str(g("emp_class")),
            "reason": _emp_id_to_str(g("reason")),
        })
    return out


def compute_withdrawal_rates(
    retiree_records: List[Dict], active_ages: List[int], obs_years: float,
    retirement_age: int = 60, band_width: int = 5,
) -> Dict:
    """직전 3~5년 퇴직자명부(분자) + 현재 재직자 연령분포(분모)로 경험 퇴직률 산출.

    중심탈퇴율 근사: q_band = 퇴직건수(밴드) / [재직자수(밴드) × 관측연수].
    사망률·승급률은 여기서 산출하지 않는다(사망률=개발원 차용, 승급률=임금인상율 입력).

    반환: {'bands':[{밴드,퇴직건수,재직자수,노출(인년),경험퇴직률}], 'rows':[{age,withdrawal,...}],
           'summary':{n_retirees,n_withdrawals,exposure,overall_withdrawal,obs_years,skipped}}
    """
    lo, hi = 15, int(retirement_age)
    bands = list(range(lo, hi, band_width)) or [lo]
    active_cnt = {b: 0 for b in bands}
    wd = {b: 0 for b in bands}

    for a in active_ages:
        if a is None:
            continue
        active_cnt[_band_of(int(a), band_width, lo, hi)] += 1

    skipped = 0
    for r in retiree_records:
        birth, leave = r.get("birth"), r.get("leave")
        if birth is None or leave is None:
            skipped += 1
            continue
        age_leave = int((leave.toordinal() - birth.toordinal()) // 365.25)
        wd[_band_of(age_leave, band_width, lo, hi)] += 1

    oy = float(obs_years) if obs_years and obs_years > 0 else 1.0
    q_by_band = {}
    band_rows = []
    tot_wd = 0
    tot_exp = 0.0
    for b in bands:
        expo = active_cnt[b] * oy
        q = (wd[b] / expo) if expo > 0 else None
        q_by_band[b] = q
        tot_wd += wd[b]
        tot_exp += expo
        top = min(b + band_width - 1, hi)
        band_rows.append({
            "밴드": f"{b}-{top}", "퇴직건수": wd[b], "재직자수": active_cnt[b],
            "노출(인년)": round(expo, 1),
            "경험퇴직률": (round(q, 5) if q is not None else None),
        })

    rows = []
    for age in range(lo, hi + 1):
        q = q_by_band.get(_band_of(age, band_width, lo, hi))
        rows.append({"age": age, "withdrawal": (round(q, 6) if q is not None else None),
                     "raise_rate": None, "mort_m": None, "mort_f": None})

    summary = {
        "n_retirees": len(retiree_records), "n_withdrawals": tot_wd,
        "exposure": round(tot_exp, 1),
        "overall_withdrawal": (tot_wd / tot_exp) if tot_exp > 0 else None,
        "obs_years": oy, "skipped": skipped,
    }
    return {"bands": band_rows, "rows": rows, "summary": summary}


def compute_experience_rates(
    records: List[Dict], obs_start: Optional[date], obs_end: Optional[date],
    retirement_age: int = 60, band_width: int = 5, min_salary_years: float = 0.5,
) -> Dict:
    """경험 퇴직률·승급률을 밴드별로 산출하고 연령(15~정년) 행으로 펼친다.

    반환: {
      'bands': [{'band':'20-24','exposure':..,'withdrawals':..,'withdrawal':q,'raise_rate':g,'raise_n':n}],
      'rows': [{'age':15,'withdrawal':q,'raise_rate':g,'mort_m':None,'mort_f':None}, ...],
      'summary': {'n':..,'n_withdrawals':..,'exposure':..,'overall_withdrawal':..,
                  'avg_raise':..,'obs_years':..,'skipped':..},
    }
    """
    lo = 15
    hi = int(retirement_age)
    bands = list(range(lo, hi, band_width)) or [lo]
    exp = {b: 0.0 for b in bands}
    wd = {b: 0 for b in bands}
    rsum = {b: 0.0 for b in bands}
    rn = {b: 0 for b in bands}
    tot_exp = 0.0
    tot_wd = 0
    skipped = 0

    for rec in records:
        birth, hire = rec.get("birth"), rec.get("hire")
        if birth is None:
            skipped += 1
            continue
        ex_start = hire if (hire and obs_start and hire > obs_start) else obs_start
        if ex_start is None:
            ex_start = hire
        leave = rec.get("leave")
        left = rec.get("left")
        ex_end = leave if (left and leave) else obs_end
        if ex_start is None or ex_end is None or ex_end <= ex_start:
            skipped += 1
            continue
        yrs = (ex_end - ex_start).days / 365.25
        if yrs <= 0:
            skipped += 1
            continue
        mid_days = ex_start.toordinal() + (ex_end - ex_start).days / 2
        age_mid = int((mid_days - birth.toordinal()) // 365.25)
        b = _band_of(age_mid, band_width, lo, hi)
        exp[b] += yrs
        tot_exp += yrs
        if left and leave and (obs_start is None or leave >= obs_start) and leave <= ex_end:
            wd[b] += 1
            tot_wd += 1
        # 승급률: 재직 지속자의 연환산 급여상승(관측 전기간 재직 표본)
        ss, se = rec.get("sal_start"), rec.get("sal_end")
        if (not left) and ss and se and ss > 0 and se > 0 and yrs >= min_salary_years:
            ann = (se / ss) ** (1.0 / yrs) - 1.0
            if -0.5 < ann < 1.0:      # 이상치 제외
                rsum[b] += ann
                rn[b] += 1

    def band_label(b):
        top = min(b + band_width - 1, hi)
        return f"{b}-{top}"

    band_rows = []
    q_by_band, g_by_band = {}, {}
    for b in bands:
        q = (wd[b] / exp[b]) if exp[b] > 0 else None
        g = (rsum[b] / rn[b]) if rn[b] > 0 else None
        q_by_band[b] = q
        g_by_band[b] = g
        band_rows.append({
            "밴드": band_label(b), "노출(인년)": round(exp[b], 2),
            "퇴직건수": wd[b], "경험퇴직률": (round(q, 5) if q is not None else None),
            "승급표본": rn[b], "경험승급률": (round(g, 5) if g is not None else None),
        })

    # 빈 밴드는 인접 밴드값으로 채움(앞→뒤, 뒤→앞)
    def _fill(d):
        keys = bands
        last = None
        for k in keys:
            if d[k] is None and last is not None:
                d[k] = last
            elif d[k] is not None:
                last = d[k]
        last = None
        for k in reversed(keys):
            if d[k] is None and last is not None:
                d[k] = last
            elif d[k] is not None:
                last = d[k]
        return d

    _fill(q_by_band)
    _fill(g_by_band)

    rows = []
    for age in range(lo, hi + 1):
        b = _band_of(age, band_width, lo, hi)
        rows.append({
            "age": age,
            "withdrawal": q_by_band.get(b) or 0.0,
            "raise_rate": g_by_band.get(b),
            "mort_m": None, "mort_f": None,
        })

    raises = [g for g in g_by_band.values() if g is not None]
    summary = {
        "n": len(records), "n_withdrawals": tot_wd, "exposure": round(tot_exp, 2),
        "overall_withdrawal": (round(tot_wd / tot_exp, 5) if tot_exp > 0 else None),
        "avg_raise": (round(sum(raises) / len(raises), 5) if raises else None),
        "obs_years": (round((obs_end - obs_start).days / 365.25, 2)
                      if (obs_start and obs_end) else None),
        "skipped": skipped,
    }
    return {"bands": band_rows, "rows": rows, "summary": summary}


def apply_mortality_from(rows: List[Dict], mort_rows: List[Dict]) -> List[Dict]:
    """경험 기초율 행에 개발원(또는 다른 세트)의 연령별 사망률(남/여)을 채운다."""
    m = {int(r["age"]): r for r in mort_rows if r.get("age") is not None}
    for r in rows:
        src = m.get(int(r["age"]))
        if src:
            r["mort_m"] = src.get("mort_m")
            r["mort_f"] = src.get("mort_f")
    return rows


def build_experience_template() -> bytes:
    """경험기초율 산출데이터 표준 양식(작성요령 + 관측기간 + 데이터 헤더) xlsx 생성."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "작성요령"
    ws.column_dimensions["A"].width = 96
    ws["A1"] = "[경험기초율 산출데이터] 작성요령"
    ws["A1"].font = Font(name="맑은 고딕", size=13, bold=True, color="1F4E79")
    guide = [
        "· 회사의 실제 퇴직·급여상승 경험으로 '경험기초율(경험 퇴직률·승급률)'을 산출하기 위한 자료입니다.",
        "· '데이터' 시트 상단의 관측시작일·관측종료일을 먼저 채우세요(예: 최근 3~5년 구간, yyyymmdd).",
        "· 관측기간 동안 재직했던 종업원(기간 중 퇴직자 포함)을 1인 1행으로 입력합니다.",
        "· 사원번호: 회사 임의 번호 (★ 실명·주민번호 등 개인정보 금지 — 자동 삭제됩니다)",
        "· 생년월일·입사일자·퇴직일자: yyyymmdd (예: 19850101). 성별: 남 1 / 여 2.",
        "· 관측시작급여: 관측시작일(또는 입사일) 시점 월 기준급여. 관측종료급여: 관측종료일(또는 퇴직일) 시점 급여.",
        "· 상태: 관측종료일 현재 재직이면 1, 관측기간 중 퇴직했으면 2(퇴직일자 필수).",
        "· 표본이 많을수록(특히 300인 이상·3년 이상) 경험률의 신뢰도가 높아집니다.",
        "· 사망률은 회사 자료로 추정하기 어려워 개발원 기초율에서 차용합니다(계리사가 세트 지정).",
    ]
    for i, line in enumerate(guide, start=3):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(name="맑은 고딕", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    ws2 = wb.create_sheet("데이터")
    ws2["A1"] = "관측시작일"; ws2["B1"] = "예: 20210101"
    ws2["A2"] = "관측종료일"; ws2["B2"] = "예: 20251231"
    cols = ["사원번호", "생년월일", "성별", "입사일자",
            "관측시작급여", "관측종료급여", "상태", "퇴직일자"]
    fill = PatternFill("solid", fgColor="DCE6F1")
    hf = Font(name="맑은 고딕", size=10, bold=True)
    for j, c in enumerate(cols, start=1):
        cell = ws2.cell(row=4, column=j, value=c)
        cell.font = hf
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws2.column_dimensions[chr(64 + j)].width = 15
    # 예시 2행
    ws2.append([]) if False else None
    ex = [
        ("A001", 19850101, 1, 20100301, 3000000, 3450000, 1, None),
        ("A002", 19900505, 2, 20180101, 2600000, 2700000, 2, 20230630),
    ]
    for i, row in enumerate(ex, start=5):
        for j, v in enumerate(row, start=1):
            if v is not None:
                ws2.cell(row=i, column=j, value=v)
    ws2.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- 연령별 퇴직률 곡선 직접 반영 (엑셀 결과값 그대로) -----------------------
_CURVE_AGE = ["연령", "나이", "age", "도달연령"]
_CURVE_RATE = ["퇴직률", "당기", "경험퇴직률", "탈퇴율", "퇴직율", "rate", "withdrawal"]


def parse_withdrawal_curve(source: Union[str, Path, bytes]) -> List[Dict]:
    """연령별 퇴직률 곡선 엑셀(2열: 연령, 퇴직률)을 그대로 읽어 기초율 행으로 변환.

    회사가 엑셀로 확정한 연령별 경험퇴직률을 산출과정 없이 '결과값'만 반영할 때 사용.
    반환: [{'age':int, 'withdrawal':float, 'raise_rate':None, 'mort_m':None, 'mort_f':None}, ...]
    - 헤더/설명 행은 무시하고, (연령=정수, 퇴직률=숫자)인 행만 채택.
    - 퇴직률이 1을 넘으면 백분율(%)로 보고 100으로 나눔(예: 5 → 0.05).
    - 작성요령 시트는 건너뛴다.
    """
    from openpyxl import load_workbook

    data = source if isinstance(source, (bytes, bytearray)) else Path(source).read_bytes()
    wb = load_workbook(io.BytesIO(bytes(data)), data_only=True)
    sheet = next((s for s in wb.sheetnames if "요령" not in s), wb.sheetnames[0])
    grid = list(wb[sheet].iter_rows(values_only=True))

    # 헤더로 연령/퇴직률 열 위치를 잡되, 없으면 앞 두 열(A=연령, B=퇴직률)로 가정.
    a_col, r_col = 0, 1
    for row in grid[:5]:
        norm = {}
        for idx, cell in enumerate(row or []):
            if cell is not None:
                norm.setdefault(_norm_header(cell), idx)
        ai = next((norm[_norm_header(a)] for a in _CURVE_AGE if _norm_header(a) in norm), None)
        ri = next((norm[_norm_header(a)] for a in _CURVE_RATE if _norm_header(a) in norm), None)
        if ai is not None and ri is not None:
            a_col, r_col = ai, ri
            break

    rows, seen = [], set()
    for row in grid:
        if not row or len(row) <= max(a_col, r_col):
            continue
        age = row[a_col]
        rate = _to_float(row[r_col])
        try:
            age = int(float(age))
        except (TypeError, ValueError):
            continue                       # 헤더('연령','당기' 등)·빈 행은 건너뜀
        if rate is None or not (0 <= age <= 120) or age in seen:
            continue
        if rate > 1:                       # 백분율 입력 보정
            rate = rate / 100.0
        seen.add(age)
        rows.append({"age": age, "withdrawal": round(rate, 6),
                     "raise_rate": None, "mort_m": None, "mort_f": None})
    rows.sort(key=lambda r: r["age"])
    return rows


def build_withdrawal_curve_template(lo: int = 15, hi: int = 70) -> bytes:
    """연령별 퇴직률 곡선 표준 양식 — 업로드 파일 포맷(연령 15~70 · '당기' 퇴직률) 그대로.

    A열=연령(lo~hi 한 행씩), B열='당기'(연령별 당기 퇴직률). 계리사가 값을 채워 올린다.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "작성요령"
    ws.column_dimensions["A"].width = 92
    ws["A1"] = "[연령별 퇴직률 곡선] 작성요령"
    ws["A1"].font = Font(name="맑은 고딕", size=13, bold=True, color="1F4E79")
    for i, line in enumerate([
        "· 회사가 엑셀 등으로 확정한 '연령별 당기 퇴직률'을 결과값 그대로 반영할 때 사용합니다.",
        "· '퇴직률' 시트: A열=연령, B열='당기' 퇴직률. 연령마다 한 행씩 값을 채우세요.",
        "· 퇴직률은 소수(0.05) 또는 %(5) 모두 인식합니다.",
        "· 빠진 연령이 있으면 인접값으로 근사될 수 있으니 정년까지 모두 채우기를 권장합니다.",
        "· 사망률·승급률은 이 곡선에 포함되지 않습니다(사망률=개발원 차용, 승급률=임금인상율 입력).",
    ], start=3):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(name="맑은 고딕", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    ws2 = wb.create_sheet("퇴직률")
    fill = PatternFill("solid", fgColor="DCE6F1")
    hf = Font(name="맑은 고딕", size=10, bold=True)
    for j, name in enumerate(["연령", "당기"], start=1):
        cell = ws2.cell(row=1, column=j, value=name)
        cell.font = hf; cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        ws2.column_dimensions[chr(64 + j)].width = 14
    for i, age in enumerate(range(lo, hi + 1), start=2):
        ws2.cell(row=i, column=1, value=age)
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
