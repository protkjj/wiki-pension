"""K-IFRS 제1019호 종업원급여 계리평가보고서 서식 생성.

고객 '사업보고서' 양식을 충실히 재현한다:
  표지 · 목차 · 의견서 · Ⅰ.개요/방법론 · Ⅱ.주석공시사항(8개 표) ·
  Ⅳ.기타 세부내역 · 참고(용어정의·주요 수리) · 개인별명세

우리 엔진이 산출하는 값(확정급여채무 현재가치, 당기근무원가, 주요 가정, 민감도,
만기, 재직자 현황)은 실제 값으로 채운다. 사외적립자산·기여금·전기 대비 조정 등
회사 재무자료가 필요한 항목은 disclosure_inputs로 주입하며, 미제공 시 0으로 둔다.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import Config
from .engine import CensusResult, calculate_census, expected_cashflows
from .models import Gender

# ── 스타일 ─────────────────────────────────────────────────────────────────
FONT = "맑은 고딕"
_INK, _ACCENT, _MUTED = "1F2A37", "1F4E79", "6B7684"
F_TITLE = Font(name=FONT, size=22, bold=True, color=_INK)
F_COVER_SUB = Font(name=FONT, size=13, color=_ACCENT)
F_SEC = Font(name=FONT, size=14, bold=True, color="FFFFFF")
F_SUB = Font(name=FONT, size=11, bold=True, color=_INK)
F_BODY = Font(name=FONT, size=10, color=_INK)
F_MUTED = Font(name=FONT, size=9, color=_MUTED)
F_HEAD = Font(name=FONT, size=10, bold=True, color=_INK)
F_NUM = Font(name=FONT, size=10, color=_INK)
F_TOT = Font(name=FONT, size=10, bold=True, color=_INK)
FILL_SEC = PatternFill("solid", fgColor=_ACCENT)
FILL_HEAD = PatternFill("solid", fgColor="DCE6F1")
FILL_TOT = PatternFill("solid", fgColor="EEF3F8")
_thin = Side(style="thin", color="BFC9D4")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
R, L, C = (Alignment(horizontal=h, vertical="center") for h in ("right", "left", "center"))
WON, PCT = "#,##0", "0.000%"


def _sheet(wb, title):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    return ws


def _para(ws, row, lines, col=1, font=F_BODY, gap_after=1):
    for ln in lines:
        ws.cell(row=row, column=col, value=ln).font = font
        row += 1
    return row + gap_after


def _sec_title(ws, row, text, ncol=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SEC
    c.fill = FILL_SEC
    c.alignment = L
    ws.row_dimensions[row].height = 24
    return row + 2


def _kv2(ws, row, title, unit, rows, cur_hdr="당기말", prior_hdr="전기말", total_label=None):
    """항목 | 당기 | 전기 형식의 조정내역 표."""
    ws.cell(row=row, column=1, value=title).font = F_SUB
    ws.cell(row=row, column=6, value=unit).font = F_MUTED
    ws.cell(row=row, column=6).alignment = R
    row += 1
    # 헤더 (열: 1=항목, 5=당기, 6=전기)
    ws.cell(row=row, column=1, value="항목").font = F_HEAD
    ws.cell(row=row, column=1).fill = FILL_HEAD
    ws.cell(row=row, column=1).border = BORDER
    for col, h in ((5, cur_hdr), (6, prior_hdr)):
        cc = ws.cell(row=row, column=col, value=h)
        cc.font = F_HEAD; cc.fill = FILL_HEAD; cc.alignment = C; cc.border = BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    for label, cur, prior in rows:
        is_tot = total_label and label == total_label
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = F_TOT if is_tot else F_BODY
        lc.border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        for col, v in ((5, cur), (6, prior)):
            cc = ws.cell(row=row, column=col, value=v)
            cc.border = BORDER; cc.alignment = R
            cc.font = F_TOT if is_tot else F_NUM
            if isinstance(v, (int, float)):
                cc.number_format = WON
            if is_tot:
                cc.fill = FILL_TOT
        if is_tot:
            lc.fill = FILL_TOT
        row += 1
    return row + 1


def _di(d, key, default=0):
    return (d or {}).get(key, default)


def _class_stats(result: CensusResult):
    labels = {"REGULAR": "일반직", "EXECUTIVE": "임원", "CONTRACT": "계약직"}
    order = ["REGULAR", "EXECUTIVE", "CONTRACT"]
    out = []
    for k in order:
        grp = [r for r in result.results if r.emp_class == k]
        if not grp:
            continue
        n = len(grp)
        out.append((labels[k], n,
                    sum(r.attained_age for r in grp) / n,
                    sum(r.attained_service for r in grp) / n,
                    sum(r.base_salary for r in grp)))
    tn = sum(x[1] for x in out) or 1
    total = ("합계", sum(x[1] for x in out),
             sum(x[1] * x[2] for x in out) / tn,
             sum(x[1] * x[3] for x in out) / tn,
             sum(x[4] for x in out))
    return out, total


def _reprice(records, config: Config, tables, dd=0.0, sd=0.0) -> float:
    data = config.model_dump()
    data["discount_rate"] = {"flat": config.discount_rate.flat + dd}
    data["salary_increase_rate"] = {"flat": config.salary_increase_rate.flat + sd}
    return calculate_census(records, Config.model_validate(data), tables, with_detail=False).total_dbo


_BANDS = ["1년 미만", "1년 이상 2년 미만", "2년 이상 3년 미만", "3년 이상 4년 미만",
          "4년 이상 5년 미만", "5년 이상 6년 미만", "6년 이상 7년 미만", "7년 이상 8년 미만",
          "8년 이상 9년 미만", "9년 이상 10년 미만", "10년 이상 15년 미만",
          "15년 이상 20년 미만", "20년 이상"]


def _bucketize(cf: pd.DataFrame, value_col: str) -> Dict[str, float]:
    b = {name: 0.0 for name in _BANDS}
    for _, r in cf.iterrows():
        t = int(r["연도"])
        if t <= 9:
            b[_BANDS[t]] += r[value_col]      # t=1 → '1년 이상 2년 미만'
        elif t <= 14:
            b["10년 이상 15년 미만"] += r[value_col]
        elif t <= 19:
            b["15년 이상 20년 미만"] += r[value_col]
        else:
            b["20년 이상"] += r[value_col]
    return b


# ═══════════════════════════════════════════════════════════════════════════
def write_report(
    out_path,
    company: str,
    valuation_date: date,
    config: Config,
    census_result: CensusResult,
    records,
    tables,
    disclosure_inputs: Optional[dict] = None,
    plan_info: Optional[dict] = None,
    prior: Optional[dict] = None,
    valuer: str = "보험계리사",
) -> Path:
    di = disclosure_inputs or {}
    y = valuation_date.year
    dbo = round(census_result.total_dbo)
    csc = round(census_result.total_csc)
    disc = config.discount_rate.flat
    sal = config.salary_increase_rate.flat
    ret_age = config.retirement_age.default

    wb = Workbook()

    # ── 표지 ──
    ws = wb.active
    ws.title = "표지"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 15
    ws["B12"] = f"{company} 를 위한"
    ws["B12"].font = F_COVER_SUB
    ws["B14"] = "K-IFRS 제1019호"
    ws["B14"].font = Font(name=FONT, size=15, bold=True, color=_ACCENT)
    ws["B16"] = "종업원급여 계리평가보고서"
    ws["B16"].font = F_TITLE
    ws["B20"] = f"{y}년 {valuation_date.month}월 {valuation_date.day}일 기준"
    ws["B20"].font = F_SUB

    # ── 목차 ──
    ws = _sheet(wb, "목차")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 60
    ws["B3"] = "- 목    차 -"
    ws["B3"].font = F_TITLE
    toc = ["K-IFRS 제1019호 종업원급여 계리평가보고서에 대한 의견서",
           "Ⅰ-1. 개요", "Ⅰ-2. 계리평가 방법론", "Ⅱ. 주석공시사항",
           "Ⅳ. 기타 세부내역", "< 참고1. 용어정의 >", "< 참고2. 주요 수리 >"]
    r = 6
    for t in toc:
        ws.cell(row=r, column=2, value=t).font = F_SUB
        r += 2

    # ── 의견서 ──
    ws = _sheet(wb, "의견서")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100
    r = 2
    ws.cell(row=r, column=2, value="K-IFRS 제1019호 종업원급여 계리평가보고서에 대한 의견서").font = F_TITLE
    r += 2
    r = _para(ws, r, [
        f"본 평가인은 K-IFRS 제1019호 '종업원급여'를 기초로 {company}(이하 '회사')의 {y}년 {valuation_date.month}월 {valuation_date.day}일",
        "까지 현재의 확정급여채무 및 다음 회계기간 동안 인식해야 할 비용을 계리평가 하였습니다.",
    ])
    r = _para(ws, r, [
        "1. 본 보고서의 목적",
        "  ⑴ K-IFRS 제1019호에 의거 회사의 회계 결산 지원 및 평가결과에 대한 보험계리사 인증",
        "  ⑵ 회사 외부감사인의 회계감사 목적으로 열람 및 활용",
        "  ⑶ 세무조정 및 퇴직연금 부담금 산출 등 재무회계 이외의 목적으로 활용하는 것은 부적절",
    ])
    r = _para(ws, r, [
        "2. 본 보고서가 포함하는 내용",
        "  ⑴ K-IFRS 제1019호 '종업원급여' 결산과 관련한 해당 계정과목 금액 계리",
        "  ⑵ K-IFRS 제1019호 '종업원급여'에서 요구하는 주석 등 공시자료 산출",
        "  ⑶ K-IFRS 제1019호 '종업원급여' 계리 결과에 대한 보험계리사의 인증",
    ])
    r = _para(ws, r, [
        "3. 본 보고서에 활용된 기초정보",
        "  ⑴ 회사가 제공한 재직자 및 퇴직자명부 등 평가 및 가정수립에 필요한 데이터",
        "  ⑵ 회사가 제공한 퇴직급여 지급액, 사외적립자산 평가액, 기여금 납부 등 회사 재무정보",
        f"  ⑶ 본 보고서의 전기말 확정급여채무·보험수리적 가정 및 당기 당기근무원가·이자비용·이자수익 등은 전기말 평가보고서를 기초로 하였습니다.",
        "  ⑷ 본 평가인은 회사가 제공한 일체의 기초정보에 대하여 제한적 검토를 실시한 결과 중대한 오류를 발견하지 못하였으며, 다만 동 검토가 정보의 완전성을 의미하는 것은 아닙니다.",
    ])
    r = _para(ws, r, [
        "4. 평가의 타당성",
        "  ⑴ 보험수리적 가정의 불편의성 및 양립가능성 확보를 위한 조언 제공",
        "  ⑵ 일반적으로 승인되고 인정된 보험수리적 방법론 및 관행에 의하여 합리적이고 타당하게 실시",
        "  ⑶ 본 계리평가는 독립적·객관적으로 진행되었으며 회사의 의견이 결과에 영향을 미치지 아니함",
    ])
    r = _para(ws, r, [
        "5. 평가인의 자격 및 독립성",
        "  ⑴ 본 평가인은 보험업법 제182조(보험계리사) 등 관계 법령에 의거 자격 요건을 충족",
        "  ⑵ 본 평가인은 회사와 어떠한 재무적 이해관계 및 인적 특수관계가 없음",
        "  ⑶ 본 평가인은 퇴직연금 등 회사와 어떠한 사업상 이해관계가 없음",
    ])

    # ── Ⅰ. 개요 및 방법론 ──
    ws = _sheet(wb, "Ⅰ.개요·방법론")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100
    r = _sec_title(ws, 2, "Ⅰ-1. 개요")
    r = _para(ws, r, [
        f"본 평가인은 {company}(이하 '회사')의 K-IFRS 제1019호 '종업원급여'와 관련하여 보험수리적 원칙과",
        "관행에 의거해 합리적이고 타당하게 작성하였습니다. 계리평가의 완전성 제고를 위하여 다음과 같이",
        "검토·평가·분석 등의 업무를 수행하였습니다.",
        "",
        "1. 원천데이터",
        "  ⑴ 재직자명부상 퇴직급여추계액과 재계산된 추계액의 비교를 통한 재직자명부의 적정성 검토",
        "  ⑵ 입사 1년 미만 신규입사자의 재직자명부 포함 여부 확인",
        "  ⑶ 평균임금이 전년 대비 또는 회사 평균 대비 과도하게 높거나 낮은 경우 재확인 요청",
        "2. 보험수리적 가정",
        "  ⑴ 산출된 가정에 대해 과거 통계치와 비교하여 급격한 변동 여부 검토",
        "  ⑵ 가정이 회사 특성과 부합하며 기초 데이터가 최근 통계를 반영하는지 검토",
        "  ⑶ 예정사망률의 경우 보험개발원 표준 참조율 업데이트 여부 확인",
        "3. 계리 설계",
        "  ⑴ 각 제도별 지급기준·정년 등 제도 규정이 산출시스템에 정확히 반영되었는지 확인",
        "  ⑵ 할인율 등 시장가치 지표가 재무제표 보고일자에 맞춰 적용되었는지 확인",
        "  ⑶ 제도·직군별 계리내용이 각 종업원에게 제대로 적용되었는지 확인",
        "4. 평가 계산",
        "  ⑴ 보험수리적손익의 원천별 분해를 통한 확정급여채무 변동 원인 분석",
        "  ⑵ 개인별 샘플 계산 확인을 통한 산출 과정상 오류 차단",
        "  ⑶ 기타 산출 과정에서 발생할 수 있는 에러 차단",
        "",
        f"본 평가인은 {y}년 {valuation_date.month}월 {valuation_date.day}일까지 현재 회사의 K-IFRS 제1019호와 관련하여 작성된 본 보고서가",
        "충분히 신뢰할 수 있고 보고서 작성 목적에 부합한다고 판단합니다.",
    ])
    r = _sec_title(ws, r + 1, "Ⅰ-2. 계리평가 방법론 (Valuation Methodology)")
    r = _para(ws, r, [
        "본 평가인은 회사 종업원급여의 채무 및 관련 비용을 측정함에 있어 K-IFRS 제1019호 및 한국계리업무기준",
        "제3편(종업원급여)에 명시된 계리적 방법론을 준수하였습니다.",
        "",
        "1. 측정 기준",
        "  ⑴ K-IFRS 제1019호 문단57·문단67의 예측단위적립방식(Projected Unit Credit Method)을 적용",
        "  ⑵ 제도특성을 수리적으로 설계하고 보험수리적 가정을 적용하여 미래현금흐름을 산출한 후,",
        "      이를 근무기간에 따라 할당·배분·할인하여 채무 및 비용을 측정",
        "2. 부채의 예상만기 및 할인율의 결정",
        "  ⑴ 채무 만기(duration) 결정 시 문단85에 따라 급여의 예상지급시기와 예상금액을 모두 고려",
        "  ⑵ 할인율은 한국계리업무기준 제3편 문단2.6.3의 수익률곡선접근법(Yield Curve Method)을 적용",
        "3. 그 외 보험수리적 가정",
        "  ⑴ 임금상승률: 대상 통계가 충분한 경우 최소자승법으로 경험승급률을 산출한 후 물가안정목표치 및",
        "      회사 중장기 임금정책을 고려하여 결정",
        "  ⑵ 예상퇴직률: 과거 통계가 충분한 경우 과거 3년간 퇴직자 데이터로 경험퇴직률 산출",
        "  ⑶ 통계가 충분하지 않은 경우 보험개발원(KIDI) 표준 기초율 적용",
    ])

    # ── Ⅱ. 주석공시사항 ──
    ws = _sheet(wb, "Ⅱ.주석공시사항")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 6
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    r = 2
    ws.cell(row=r, column=1, value="Ⅱ. 주석공시사항").font = F_TITLE
    r += 2

    assets = _di(di, "plan_assets")
    npc = _di(di, "npc_conversion")
    ceiling = _di(di, "asset_ceiling")
    net = round(-assets + dbo + ceiling + npc) if (assets or npc) else dbo
    r = _para(ws, r, ["1. 전기말과 당기말 현재 재무상태표에 계상된 순확정급여부채(자산)의 조정내역은 다음과 같습니다."])
    r = _kv2(ws, r, "", "(단위: 원)", [
        ("사외적립자산", -assets, -_di(di, "plan_assets_prior")),
        ("확정급여채무의 현재가치", dbo, _di(di, "dbo_prior")),
        ("자산인식상한효과", ceiling, 0),
        ("국민연금전환금", -npc, -_di(di, "npc_conversion")),
        ("순확정급여부채(자산)", net, ""),
    ], total_label="순확정급여부채(자산)")

    r = _para(ws, r, ["2. 전기와 당기 중 사외적립자산의 조정내역은 다음과 같습니다."])
    r = _kv2(ws, r, "", "(단위: 원)", [
        ("기초 사외적립자산", _di(di, "plan_assets_begin"), ""),
        ("이자수익", _di(di, "interest_income"), ""),
        ("기여금납부액", _di(di, "contributions"), ""),
        ("급여지급액(중간정산금액 포함)", -_di(di, "benefits_paid"), ""),
        ("사외적립자산의 운영비용", 0, ""),
        ("사외적립자산의 수익(순이자 포함액 제외)", _di(di, "asset_return"), ""),
        ("기말 사외적립자산", assets, ""),
    ], cur_hdr=f"{y}년말", prior_hdr=f"{y-1}년", total_label="기말 사외적립자산")

    ic = round(_di(di, "dbo_begin") * disc) if _di(di, "dbo_begin") else _di(di, "interest_cost")
    r = _para(ws, r, ["3. 전기와 당기 중 확정급여채무의 현재가치 조정내역은 다음과 같습니다."])
    r = _kv2(ws, r, "", "(단위: 원)", [
        ("기초 확정급여채무", _di(di, "dbo_begin"), ""),
        ("당기근무원가 주1)", csc, ""),
        ("이자비용", ic, ""),
        ("급여지급액(중간정산금액 포함)", -_di(di, "benefits_paid_dbo"), ""),
        ("인구통계적가정 변동 보험수리적손익", _di(di, "remeasure_demographic"), ""),
        ("재무적가정 변동 보험수리적손익", _di(di, "remeasure_financial"), ""),
        ("가정과 실제 차이 보험수리적손익", _di(di, "remeasure_experience"), ""),
        ("과거근무원가와 정산 손익", 0, ""),
        ("기말 확정급여채무", dbo, ""),
    ], cur_hdr=f"{y}년말", prior_hdr=f"{y-1}년", total_label="기말 확정급여채무")
    ws.cell(row=r, column=1, value="주1) 신규 입사자의 당기근무원가가 포함되어 있습니다.").font = F_MUTED
    r += 2

    net_interest = _di(di, "net_interest")
    r = _para(ws, r, ["4. 당기손익으로 인식하는 확정급여원가는 다음과 같습니다."])
    r = _kv2(ws, r, "", "(단위: 원)", [
        ("당기근무원가", csc, ""),
        ("과거근무원가와 정산 손익", 0, ""),
        ("순확정급여부채(자산)의 순이자", net_interest, ""),
        ("사외적립자산의 운영비용", 0, ""),
        ("당기손익으로 인식하는 확정급여원가", round(csc + net_interest), ""),
    ], cur_hdr=f"{y}년말", prior_hdr=f"{y-1}년", total_label="당기손익으로 인식하는 확정급여원가")

    r = _para(ws, r, ["5. 기타포괄손익으로 인식하는 확정급여원가는 다음과 같습니다. (최초 평가 시 미산출)"])
    r = _kv2(ws, r, "", "(단위: 원)", [
        ("인구통계적가정 변동 보험수리적손익", _di(di, "remeasure_demographic"), ""),
        ("재무적가정 변동 보험수리적손익", _di(di, "remeasure_financial"), ""),
        ("가정과 실제 차이 보험수리적손익", _di(di, "remeasure_experience"), ""),
        ("사외적립자산의 수익(순이자 포함액 제외)", -_di(di, "asset_return"), ""),
        ("자산인식상한효과의 변동", 0, ""),
        ("기타포괄손익으로 인식하는 확정급여원가", "", ""),
    ], cur_hdr=f"{y}년말", prior_hdr=f"{y-1}년")

    # 6. 주요 가정
    ws.cell(row=r, column=1, value="6. 전기말과 당기말 현재 사용한 주요 보험수리적 가정은 다음과 같습니다.").font = F_SUB
    r += 1
    for col, h in ((1, "가정"), (5, f"{y}년말"), (6, f"{y-1}년말")):
        cc = ws.cell(row=r, column=col, value=h); cc.font = F_HEAD; cc.fill = FILL_HEAD; cc.border = BORDER
        cc.alignment = C if col > 1 else L
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    for label, v, isp in [("할인율", disc, True), ("임금상승률(Base-up)", sal, True),
                          ("승급률", config.salary_increase_timing, False),
                          ("퇴직률", f"{config.retirement_rate_basis} 기준 참조율", False),
                          ("사망률", "성별·연령별 참조율", False)]:
        lc = ws.cell(row=r, column=1, value=label); lc.font = F_BODY; lc.border = BORDER
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        vc = ws.cell(row=r, column=5, value=v); vc.font = F_NUM; vc.border = BORDER; vc.alignment = R
        if isp:
            vc.number_format = PCT
        ws.cell(row=r, column=6).border = BORDER
        r += 1
    r += 1

    # 7. 민감도 (±1%p)
    d_up, d_dn = _reprice(records, config, tables, dd=+0.01), _reprice(records, config, tables, dd=-0.01)
    s_up, s_dn = _reprice(records, config, tables, sd=+0.01), _reprice(records, config, tables, sd=-0.01)
    ws.cell(row=r, column=1, value="7. 주요 보험수리적 가정에 대한 확정급여채무의 변동 (±1.00%p)").font = F_SUB
    r += 1

    def sens_block(row, title, base, up, dn):
        ws.cell(row=row, column=1, value=title).font = F_BODY
        row += 1
        for col, h in ((4, "1.00%p 상승"), (5, "기준"), (6, "1.00%p 하락")):
            cc = ws.cell(row=row, column=col, value=h); cc.font = F_HEAD; cc.fill = FILL_HEAD; cc.border = BORDER; cc.alignment = C
        row += 1
        for label, vals in [("확정급여채무", (up, base, dn)),
                            ("확정급여채무의 변화량", (up - base, 0, dn - base))]:
            ws.cell(row=row, column=1, value=label).font = F_BODY
            ws.cell(row=row, column=1).border = BORDER
            for col, v in zip((4, 5, 6), vals):
                cc = ws.cell(row=row, column=col, value=round(v)); cc.font = F_NUM; cc.border = BORDER; cc.alignment = R; cc.number_format = WON
            row += 1
        # 기말채무 대비 비율
        ws.cell(row=row, column=1, value="기말채무 대비 비율").font = F_BODY
        ws.cell(row=row, column=1).border = BORDER
        for col, v in zip((4, 5, 6), (up / base, 1.0, dn / base)):
            cc = ws.cell(row=row, column=col, value=v); cc.font = F_NUM; cc.border = BORDER; cc.alignment = R; cc.number_format = "0.0000"
        return row + 2

    r = sens_block(r, "(1) 할인율의 변동", dbo, d_up, d_dn)
    r = sens_block(r, "(2) 임금상승률의 변동(Base-up)", dbo, s_up, s_dn)

    # 8. 만기구성정보
    cf = expected_cashflows(records, config, tables)
    pv_b = _bucketize(cf, "현재가치")
    pay_b = _bucketize(cf, "기대급여지급액")
    dur = (cf["연도"] * cf["현재가치"]).sum() / cf["현재가치"].sum() if len(cf) and cf["현재가치"].sum() else 0
    ws.cell(row=r, column=1, value="8. 당기말 확정급여채무의 만기구성정보").font = F_SUB
    r += 1
    ws.cell(row=r, column=2, value="(1) 만기별 확정급여채무").font = F_BODY
    r += 1
    for col, h in ((2, "만기"), (5, "확정급여채무")):
        cc = ws.cell(row=r, column=col, value=h); cc.font = F_HEAD; cc.fill = FILL_HEAD; cc.border = BORDER; cc.alignment = C if col > 2 else L
    r += 1
    for name in _BANDS:
        ws.cell(row=r, column=2, value=name).font = F_BODY
        ws.cell(row=r, column=2).border = BORDER
        cc = ws.cell(row=r, column=5, value=round(pv_b[name])); cc.font = F_NUM; cc.border = BORDER; cc.alignment = R; cc.number_format = WON
        r += 1
    # 간편법(제도2) 평가분은 PUC 만기배분 대상이 아니므로 별도 조정행으로 표시해 총계를 재무상태표 DBO에 일치시킨다.
    simple_dbo = census_result.subtotal_by_plan.get(2, {}).get("DBO", 0.0)
    if simple_dbo:
        ws.cell(row=r, column=2, value="간편법 평가분(제도2)").font = F_BODY
        ws.cell(row=r, column=2).border = BORDER
        cc = ws.cell(row=r, column=5, value=round(simple_dbo)); cc.font = F_NUM; cc.border = BORDER; cc.alignment = R; cc.number_format = WON
        r += 1
    ws.cell(row=r, column=2, value="합계").font = F_TOT
    cc = ws.cell(row=r, column=5, value=dbo); cc.font = F_TOT; cc.number_format = WON; cc.alignment = R; cc.fill = FILL_TOT
    r += 1
    ws.cell(row=r, column=1, value=f"주1) 당기말 확정급여채무의 가중평균만기는 {dur:.2f}년입니다.").font = F_MUTED
    r += 1
    ws.cell(row=r, column=1, value="주2) 만기는 개인별 기대잔존근무년수에 따라 구분하였습니다.").font = F_MUTED
    r += 1
    if simple_dbo:
        ws.cell(row=r, column=1, value="주3) 간편법(제도2) 평가분은 PUC 만기배분 대상이 아니어서 별도 표시하였으며, 총계는 재무상태표상 확정급여채무와 일치합니다.").font = F_MUTED
        r += 1
    r += 1
    ws.cell(row=r, column=2, value="(2) 향후 예상 퇴직급여지급액").font = F_BODY
    r += 1
    for col, h in ((2, "급여지급시기"), (5, "예상 퇴직급여지급액")):
        cc = ws.cell(row=r, column=col, value=h); cc.font = F_HEAD; cc.fill = FILL_HEAD; cc.border = BORDER; cc.alignment = C if col > 2 else L
    r += 1
    for name in _BANDS:
        ws.cell(row=r, column=2, value=name).font = F_BODY
        ws.cell(row=r, column=2).border = BORDER
        cc = ws.cell(row=r, column=5, value=round(pay_b[name])); cc.font = F_NUM; cc.border = BORDER; cc.alignment = R; cc.number_format = WON
        r += 1

    # ── Ⅳ. 기타 세부내역 ──
    ws = _sheet(wb, "Ⅳ.기타세부내역")
    ws.column_dimensions["A"].width = 3
    for col, w in zip("BCDEF", (26, 14, 14, 18, 18)):
        ws.column_dimensions[col].width = w
    r = 2
    ws.cell(row=r, column=1, value="Ⅳ. 기타 세부내역").font = F_TITLE
    r += 2
    r = _para(ws, r, [
        "1. 확정급여채무의 보험수리적 평가는 예측단위적립방식(PUC)을 사용하였습니다.",
        "   임원·계약직·정년초과자의 경우 회사가 제공한 금액을 기준으로 평가방식을 준용하였습니다.",
    ])
    plan_line1 = f"종업원의 정년은 만 {ret_age}세 입니다."
    if plan_info:
        pb = plan_info.get("benefit_rule") or "-"
        sb = plan_info.get("salary_basis") or "평균임금"
        plan_line2 = f"퇴직 시 근속 1년당 {sb} 기준으로 산정한 금액을 퇴직금으로 지급합니다({pb})."
    else:
        plan_line2 = "퇴직 시 근속 1년당 최종 평균임금에 지급률을 곱한 금액을 퇴직금으로 지급합니다."
    r = _para(ws, r, [
        "2. 퇴직금제도 요약",
        plan_line1,
        plan_line2 + " 퇴직금은 일시금으로 지급하며, 근속 1년 미만 기간은 일할 계산하였습니다.",
    ])

    # 3. 재직자 현황
    stats, total = _class_stats(census_result)
    ws.cell(row=r, column=1, value=f"3. {valuation_date} 현재 재직자에 대하여 확정급여채무를 평가하였으며, 현황은 다음과 같습니다.").font = F_SUB
    r += 1
    heads = ["구분", "인원(명)", "평균연령(세)", "평균근속년수(년)", "평균임금 합(원)"]
    for j, h in enumerate(heads):
        cc = ws.cell(row=r, column=2 + j, value=h); cc.font = F_HEAD; cc.fill = FILL_HEAD; cc.border = BORDER; cc.alignment = C
    r += 1
    for name, n, aage, asvc, ssal in stats + [total]:
        istot = name == "합계"
        vals = [name, n, round(aage, 2), round(asvc, 2), round(ssal)]
        for j, v in enumerate(vals):
            cc = ws.cell(row=r, column=2 + j, value=v)
            cc.font = F_TOT if istot else (F_BODY if j == 0 else F_NUM)
            cc.border = BORDER; cc.alignment = L if j == 0 else R
            if j == 4:
                cc.number_format = WON
            if istot:
                cc.fill = FILL_TOT
        r += 1
    r += 1

    # 4. 가정 세부 + 예시연령별 기초율
    r = _para(ws, r, ["4. 보험수리적 가정의 세부내역", "(1) 재무적 가정"])
    for label, v, fmt in [("할인율", disc, PCT), ("임금상승률(Base-up)", sal, PCT),
                          ("기대잔존근무년수(년)", round(dur, 2), None)]:
        ws.cell(row=r, column=2, value=label).font = F_BODY
        cc = ws.cell(row=r, column=3, value=v); cc.font = F_NUM; cc.alignment = R
        if fmt:
            cc.number_format = fmt
        r += 1
    r += 1
    r = _para(ws, r, ["(2) 인구통계적 가정 · 예시연령별 기초율"])
    for col, h in ((2, "예시연령"), (3, "퇴직률"), (4, "사망률(남)"), (5, "사망률(여)")):
        cc = ws.cell(row=r, column=col, value=h); cc.font = F_HEAD; cc.fill = FILL_HEAD; cc.border = BORDER; cc.alignment = C
    r += 1
    for age in (20, 25, 30, 35, 40, 45, 50, 55):
        try:
            q_ret = tables.retirement_rate_by_age(age)
        except Exception:
            q_ret = None
        qm = tables.mortality_rate(age, Gender.M)
        qf = tables.mortality_rate(age, Gender.F)
        for col, v in ((2, age), (3, q_ret), (4, qm), (5, qf)):
            cc = ws.cell(row=r, column=col, value=v); cc.font = F_NUM; cc.border = BORDER; cc.alignment = R
            if col > 2 and isinstance(v, float):
                cc.number_format = "0.00000"
        r += 1

    # ── 참고 ──
    ws = _sheet(wb, "참고")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100
    r = _para(ws, 2, [
        "< 참고1. 용어정의 >",
        "* 종업원급여: 종업원이 제공한 근무용역과 교환하여 기업이 제공하는 모든 종류의 대가",
        "* 퇴직급여: 퇴직 이후에 지급하는 종업원급여(해고급여·단기종업원급여 제외)",
        "* 확정급여제도: 확정기여제도 이외의 모든 퇴직급여제도",
        "* 확정급여채무의 현재가치: 종업원이 당기·과거기간 근무용역을 제공하여 발생한 채무를 결제하는 데",
        "   필요한 예상 미래지급액의 현재가치(사외적립자산 차감 전)",
        "* 당기근무원가: 당기에 근무용역을 제공하여 발생한 확정급여채무 현재가치의 증가",
        "* 순확정급여부채(자산)의 순이자: 시간의 경과에 따라 발생하는 순확정급여부채(자산)의 변동",
        "* 보험수리적손익(재측정손익): 경험조정 및 보험수리적 가정 변경효과로 인한 확정급여채무 현재가치의",
        "   변동으로, 기타포괄손익(OCI)에 반영",
    ], col=2)
    _para(ws, r, [
        "< 참고2. 주요 수리 >",
        "* 미래급여 예측: 미래급여 = 현재급여 × (1 + 임금상승률)^t",
        "* 근무기간 배분: 근무비율 = 현재 근무기간 ÷ 총 근무기간",
        "* 확정급여채무: DBO = Σ (퇴직급여 × 근무비율 × 퇴직확률 × 할인계수)",
        "   확정급여채무 변동구조 = 기초 PBO + 당기근무원가 + 이자원가 − 지급액 ± 재측정요소 = 기말 PBO",
        "* 할인율: 현재가치(PV) = 미래급여 ÷ (1 + r)^t, 수익률곡선접근법으로 계산",
        "* 가중평균만기(듀레이션) = Σ (PBO × 지급시기) ÷ Σ PBO",
        "* 이자원가 = 기초 DBO × 할인율 − 예상지급액 × 적용할인율",
    ], col=2)

    # ── 개인별명세 ──
    ws = _sheet(wb, "개인별명세")
    df = census_result.to_dataframe()
    heads = ["사번", "종업원구분", "제도구분", "도달연령", "도달근속", "기준급여", "당년도추계액", "DBO", "CSC"]
    for j, h in enumerate(heads):
        cc = ws.cell(row=1, column=1 + j, value=h); cc.font = F_HEAD; cc.fill = FILL_HEAD; cc.border = BORDER; cc.alignment = C
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        vals = [row["emp_id"], row["emp_class"], row["plan_type"], round(row["attained_age"], 1),
                round(row["attained_service"], 1), row["base_salary"], row["current_year_accrual"],
                row["DBO"], row["CSC"]]
        for j, v in enumerate(vals):
            cc = ws.cell(row=i, column=1 + j, value=v); cc.border = BORDER; cc.font = F_NUM
            if j >= 5 and isinstance(v, (int, float)):
                cc.number_format = WON; cc.alignment = R
    for j, w in enumerate([12, 12, 12, 10, 10, 14, 16, 16, 14]):
        ws.column_dimensions[get_column_letter(1 + j)].width = w
    ws.freeze_panes = "A2"

    wb.save(Path(out_path))
    return Path(out_path)
