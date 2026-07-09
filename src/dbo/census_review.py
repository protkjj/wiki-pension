"""명부 검토 워크북 생성 — 표준변환 · 오류검토 · 안내.

기업이 올린 명부를 우리 표준양식으로 변환한 결과와, 오류로 추정되는 행을
'검토용' 시트로 묶어 한 파일로 돌려준다. 기업은 이 파일을 받아 ① 값이 맞는지
확인하고 ② 오류 추정 행을 고쳐서 다시 올리거나, 고치지 않을 경우 사유를 적어
제출한다. 명부 인테이크의 왕복·민원을 줄이기 위한 도구.
"""

from __future__ import annotations

import io
from typing import List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

# 표준 스키마(영문) → 기업이 이해하는 한글 라벨
STANDARD_KO = {
    "emp_id": "사원번호",
    "birth_date": "생년월일",
    "gender": "성별",
    "hire_date": "입사일자",
    "base_salary": "기준급여",
    "current_year_accrual": "당년도퇴직금추계액",
    "next_year_accrual": "차년도퇴직금추계액",
    "emp_class": "종업원구분",
    "interim_settlement_date": "중간정산일",
    "interim_settlement_amount": "중간정산액",
    "plan_type": "제도구분",
    "multiplier": "적용배수",
    "ifrs_enrolled": "IFRS가입",
}
_ANNOTATE_COLS = ["회사수정값(맞게 고쳐 적기)", "오류아님_사유(수정 안 할 때)"]


def standard_view(mapped_df: pd.DataFrame) -> pd.DataFrame:
    """매핑된(표준 영문컬럼) 명부를 한글 표준양식으로 정리 — 표준 컬럼만, 순서 고정."""
    cols = [c for c in STANDARD_KO if c in mapped_df.columns]
    view = mapped_df[cols].copy()
    view.columns = [STANDARD_KO[c] for c in cols]
    return view


ANNOT_COLS = ["오류가 아닐시 사유 기록", "구분", "추정오류"]


def annotate_original(raw_df: pd.DataFrame, row_issues: dict) -> pd.DataFrame:
    """기업이 올린 원본 양식(컬럼 그대로)에 추정오류를 덧붙인다.

    row_issues: {행인덱스(0-based): {"구분": "오류"|"경고", "추정오류": "..."}}
    반환: 원본 컬럼 + [오류가 아닐시 사유 기록(빈칸), 구분, 추정오류].
    원본을 그대로 두므로 기업이 바로 고쳐서 다시 올릴 수 있다.
    """
    out = raw_df.copy().reset_index(drop=True)
    out["오류가 아닐시 사유 기록"] = ""
    out["구분"] = [row_issues.get(i, {}).get("구분", "") for i in range(len(out))]
    out["추정오류"] = [row_issues.get(i, {}).get("추정오류", "") for i in range(len(out))]
    return out


def _guide_rows(dropped_pii: List[str], summary: dict) -> pd.DataFrame:
    lines = [
        "【 이 파일 사용법 】",
        "1) '①표준변환(우리양식)' 시트: 회사가 올린 명부를 우리 표준양식으로 변환한 결과입니다. 값이 맞는지 확인하세요.",
        "2) '②오류검토' 시트: 오류로 추정되는 행만 모았습니다. ⚠문제 열의 사유를 보고",
        "   · 잘못된 값이면 '회사수정값' 열에 바르게 적어 고친 뒤 명부를 다시 올려주세요.",
        "   · 오류가 아니라면 '오류아님_사유' 열에 이유를 적어주시면 그대로 접수됩니다.",
        "3) 개인정보(성명·주민번호 등)는 자동 삭제되어 등록됩니다. 사원번호만 사용합니다.",
        "",
        "【 요약 】",
        f"· 파일명: {summary.get('filename', '-')}",
        f"· 산출기준일: {summary.get('valuation_date', '-')}",
        f"· 인원(행): {summary.get('records', 0)}명",
        f"· 오류 추정: {summary.get('errors', 0)}건 · 경고: {summary.get('warnings', 0)}건",
    ]
    if dropped_pii:
        lines += ["", "【 자동 삭제된 개인정보 컬럼 】", "· " + ", ".join(dropped_pii)]
    return pd.DataFrame({"안내": lines})


def build_review_workbook(mapped_df: pd.DataFrame, problem_df: Optional[pd.DataFrame],
                          issues_df: Optional[pd.DataFrame],
                          dropped_pii: List[str], summary: dict) -> bytes:
    """검토 워크북(①표준변환 ②오류검토 ③오류목록 ④안내) xlsx 바이트 생성."""
    std = standard_view(mapped_df) if mapped_df is not None and not mapped_df.empty \
        else pd.DataFrame({"안내": ["표준변환 결과 없음"]})

    pdf = problem_df.copy() if problem_df is not None and not problem_df.empty else pd.DataFrame()
    if not pdf.empty:
        for c in _ANNOTATE_COLS:
            pdf[c] = ""
    else:
        pdf = pd.DataFrame({"안내": ["오류로 추정되는 행이 없습니다 ✅"]})

    idf = issues_df if issues_df is not None and not issues_df.empty \
        else pd.DataFrame({"안내": ["오류/경고 없음 ✅"]})

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        std.to_excel(w, sheet_name="①표준변환(우리양식)", index=False)
        pdf.to_excel(w, sheet_name="②오류검토", index=False)
        idf.to_excel(w, sheet_name="③오류목록", index=False)
        _guide_rows(dropped_pii, summary).to_excel(w, sheet_name="④안내", index=False)
        _style(w)
    return buf.getvalue()


def _style(writer) -> None:
    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")
    for name, ws in writer.sheets.items():
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # 열 너비
        for col in ws.columns:
            letter = col[0].column_letter
            width = 14
            header = str(col[0].value or "")
            if "안내" in header or "사유" in header or "문제" in header or "수정값" in header:
                width = 42
            ws.column_dimensions[letter].width = width
        ws.freeze_panes = "A2"
        # 오류검토 시트의 회사기입 열 강조
        if name == "②오류검토":
            for cell in ws[1]:
                if str(cell.value) in _ANNOTATE_COLS:
                    cell.fill = warn_fill
                    cell.font = Font(name="맑은 고딕", size=10, bold=True, color="833C00")
