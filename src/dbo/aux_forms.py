"""보조 입력 양식 3종 — 사외적립자산 · 기타장기 · 명부확인용요약표.

기업 담당자가 하나하나 화면에 입력하는 번거로움을 줄이기 위해,
각 항목을 ① 양식(xlsx) 다운로드 → ② 작성 후 업로드 → ③ 결과값만 화면에서
편집 미리보기 하는 흐름으로 처리한다.

- build_*_template(): openpyxl로 양식(녹색 입력칸)을 생성한다.
- parse_*_upload(): 업로드 파일에서 입력값(라벨 기준 스캔)을 읽어 dict/rows로 반환.
- compute_census_summary(): 재직자명부 레코드로 요약표 '자동산출' 값을 계산한다.

명부확인용요약표는 재직자명부 등록 이후에만 의미가 있으며, 양식 다운로드 시
'등록부명에서 자동산출된 값'(G열)을 재직자명부에서 산출해 채워 준다.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Dict, List, Optional, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# 사용자 제공 양식 원본(시인성 좋은 실제 양식) — 3개 시트를 그대로 다운로드에 사용한다.
ASSET_PATH = Path(__file__).resolve().parent / "assets" / "aux_templates.xlsx"
_SHEET_FUNDING = "사외적립자산"
_SHEET_OTHER_LT = "기타 장기"
_SHEET_SUMMARY = "명부확인용요약표"

# 명부확인용요약표 G열(자동산출) 셀 ↔ compute_census_summary 키 매핑(사용자 양식 좌표).
_SUMMARY_G_CELLS = {
    "G6": "재직_임원", "G7": "재직_직원", "G8": "재직_계약직", "G9": "재직_합계",
    "G14": "중간정산자수",
    "G16": "추계액_임원", "G17": "추계액_직원", "G18": "추계액_계약직", "G19": "추계액_합계",
    "G22": "중간정산금액",
}


def _extract_sheet(sheet_name: str):
    """원본 양식 워크북에서 지정 시트만 남긴 워크북을 반환(스타일·서식 보존)."""
    wb = load_workbook(ASSET_PATH)
    for s in list(wb.sheetnames):
        if s != sheet_name:
            del wb[s]
    return wb


def _to_bytes(wb) -> bytes:
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ===========================================================================
# 할인율(회사채 등급·만기 공시금리) 업로드 양식
# ===========================================================================
DISCOUNT_RATINGS = ["AAA", "AA+", "AA0", "AA-", "A+", "A0"]


def _rate_val(v):
    """'3.16%' · '3.16' · 0.0316 → 소수(0.0316). 1보다 크면 %로 간주."""
    if v is None or v == "":
        return None
    try:
        x = float(str(v).replace("%", "").replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    if x <= 0:
        return None
    return x / 100 if x > 1 else x


def build_discount_upload_template(max_maturity: int = 20) -> bytes:
    """할인율 관리 업로드 양식 — 기준일 + 등급별(만기·할인율) 와이드 표."""
    wb = Workbook()
    ws = wb.active
    ws.title = "할인율"
    ws["A1"] = "기준일(YYYYMM 또는 YYYY-MM-DD)"
    ws["A1"].font = _BOLD
    ws["B1"] = ""
    ws["B1"].fill = _GREEN
    ws["A1"].alignment = _LEFT
    ws.column_dimensions["A"].width = 30

    for idx, rating in enumerate(DISCOUNT_RATINGS):
        c0 = 1 + idx * 3            # 등급마다 3칸(만기·할인율·공백) 간격
        hdr = ws.cell(row=3, column=c0, value=f"회사채 {rating}")
        hdr.font = _BOLD; hdr.fill = _HEAD; hdr.alignment = _CEN
        ws.merge_cells(start_row=3, start_column=c0, end_row=3, end_column=c0 + 1)
        m = ws.cell(row=4, column=c0, value="만기"); m.font = _BOLD; m.fill = _HEAD; m.alignment = _CEN
        r = ws.cell(row=4, column=c0 + 1, value="할인율"); r.font = _BOLD; r.fill = _HEAD; r.alignment = _CEN
        ws.column_dimensions[chr(64 + c0)].width = 8
        ws.column_dimensions[chr(64 + c0 + 1)].width = 10
        for k in range(1, max_maturity + 1):
            ws.cell(row=4 + k, column=c0, value=k).alignment = _CEN
            gc = ws.cell(row=4 + k, column=c0 + 1); gc.fill = _GREEN
    ws.freeze_panes = "A5"
    return _to_bytes(wb)


def parse_discount_upload(source: Union[str, bytes]) -> Dict:
    """할인율 업로드(와이드) 파싱 → {'기준일': str|None, 'curves': {등급: [{maturity,rate}]}}."""
    ws = _open_first_sheet(source)
    grid = [list(r) for r in ws.iter_rows(values_only=True)]

    def _looks_like_date(v) -> Optional[str]:
        if isinstance(v, (int, float)) and 190000 < v < 99999999:
            return str(int(v))
        if isinstance(v, str):
            s = v.replace("-", "").replace(".", "").replace("/", "").strip()
            if s.isdigit() and (6 <= len(s) <= 8) and s[:4].isdigit() and 1900 < int(s[:4]) < 2100:
                return v.strip()
        return None

    base_date = None
    for row in grid[:6]:
        for j, c in enumerate(row or []):
            if c is not None and "기준일" in _norm(c):
                for v in (row[j + 1:] if j + 1 < len(row) else []):
                    if v not in (None, ""):
                        base_date = str(v).strip()
                        break
            # 값이 라벨 없이 202512 · 2025-12-31 같은 형태로만 있을 때
            if base_date is None:
                cand = _looks_like_date(c)
                if cand:
                    base_date = cand
        if base_date:
            break

    # 등급 헤더 위치 탐색 (회사채 A0 등)
    rating_norm = {_norm(r): r for r in DISCOUNT_RATINGS}
    headers = []   # (row, col, rating)
    for ridx, row in enumerate(grid[:8]):
        for cidx, c in enumerate(row or []):
            if c is None:
                continue
            n = _norm(c).replace("회사채", "")
            if n in rating_norm:
                headers.append((ridx, cidx, rating_norm[n]))

    curves: Dict[str, List[dict]] = {}
    for (hr, hc, rating) in headers:
        # 헤더 아래에서 '만기'/'할인율' 서브헤더 행 찾기
        sub_r = None
        for x in range(hr + 1, min(hr + 4, len(grid))):
            if hc < len(grid[x]) and _norm(grid[x][hc]) == "만기":
                sub_r = x
                break
        if sub_r is None:
            sub_r = hr + 1
        pts = []
        for x in range(sub_r + 1, len(grid)):
            row = grid[x]
            mv = row[hc] if hc < len(row) else None
            rv = row[hc + 1] if hc + 1 < len(row) else None
            try:
                mm = int(float(str(mv).strip()))
            except (TypeError, ValueError):
                continue
            rr = _rate_val(rv)
            if rr is not None:
                pts.append({"maturity": mm, "rate": rr})
        if pts:
            curves[rating] = sorted(pts, key=lambda p: p["maturity"])
    return {"기준일": base_date, "curves": curves}


def build_simple_curve_template(max_maturity: int = 20) -> bytes:
    """전기 할인율 업로드용 단순 양식 — 만기 · 할인율 2열."""
    wb = Workbook()
    ws = wb.active
    ws.title = "할인율"
    ws["A1"] = "만기"; ws["B1"] = "할인율"
    for cell in ("A1", "B1"):
        ws[cell].font = _BOLD; ws[cell].fill = _HEAD; ws[cell].alignment = _CEN
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    for k in range(1, max_maturity + 1):
        ws.cell(row=1 + k, column=1, value=k).alignment = _CEN
        ws.cell(row=1 + k, column=2).fill = _GREEN
    ws.freeze_panes = "A2"
    return _to_bytes(wb)


def parse_simple_curve(source: Union[str, bytes]) -> List[dict]:
    """단순 2열(만기·할인율) 업로드 → [{maturity, rate}]. 전기 할인율 업로드용."""
    ws = _open_first_sheet(source)
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    # 만기/할인율 헤더 열 찾기
    mcol = rcol = None
    hdr_row = 0
    for ridx, row in enumerate(grid[:6]):
        for cidx, c in enumerate(row or []):
            n = _norm(c)
            if n in ("만기", "만기년", "maturity"):
                mcol = cidx; hdr_row = ridx
            if n in ("할인율", "금리", "rate", "수익률"):
                rcol = cidx
        if mcol is not None and rcol is not None:
            break
    if mcol is None or rcol is None:
        mcol, rcol, hdr_row = 0, 1, 0     # 헤더 없으면 1·2열로 가정
    pts = []
    for row in grid[hdr_row + 1:]:
        mv = row[mcol] if mcol < len(row) else None
        rv = row[rcol] if rcol < len(row) else None
        try:
            mm = int(float(str(mv).strip()))
        except (TypeError, ValueError):
            continue
        rr = _rate_val(rv)
        if rr is not None:
            pts.append({"maturity": mm, "rate": rr})
    return sorted(pts, key=lambda p: p["maturity"])

# 녹색(입력) 칸 · 헤더 스타일 -------------------------------------------------
_GREEN = PatternFill("solid", fgColor="E2EFDA")     # 입력칸(녹색)
_HEAD = PatternFill("solid", fgColor="DCE6F1")       # 헤더(연파랑)
_AUTO = PatternFill("solid", fgColor="FFF2CC")       # 자동산출(연노랑)
_BOLD = Font(name="맑은 고딕", size=10, bold=True)
_NORM = Font(name="맑은 고딕", size=10)
_TITLE = Font(name="맑은 고딕", size=13, bold=True, color="1F4E79")
_CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_thin = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", "").replace("%", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _norm(s) -> str:
    """라벨 정규화 — 공백·괄호·기호 제거로 스캔 매칭 안정화."""
    if s is None:
        return ""
    t = str(s)
    for ch in " \n\t()（）[]{}·.":
        t = t.replace(ch, "")
    return t.strip()


def _open_first_sheet(source: Union[str, bytes]):
    if isinstance(source, (bytes, bytearray)):
        wb = load_workbook(io.BytesIO(source), data_only=True)
    else:
        wb = load_workbook(str(source), data_only=True)
    name = next((s for s in wb.sheetnames if "요령" not in s), wb.sheetnames[0])
    return wb[name]


# ===========================================================================
# 1) 사외적립자산 현황
# ===========================================================================
# (라벨 → funding_status data_json 키). 라벨은 _norm 적용 기준.
_FUNDING_ROWS = [
    ("기초잔액", "기초 잔액 (A)", "기초부터 이월된 기초 잔액"),
    ("입금액", "입금(수탁)액 (B)", "기초부터 누계 입금(수탁)액"),
    ("지급_퇴직", "급여지급-퇴직 (a)", "기초부터 누계 급여지급(인출)-퇴직"),
    ("지급_중간정산", "급여지급-중간정산 (b)", ""),
    ("지급_DC전환", "급여지급-DC전환 (c)", ""),
    ("관계사전입", "관계사전입 (d)", "기초부터 누계 관계사전입"),
    ("관계사전출", "관계사전출 (e)", "기초부터 누계 관계사전출"),
    ("사업결합", "사업결합 (f)", "기초부터 누계 사업결합액"),
    ("사업처분", "사업처분 (g)", "기초부터 누계 사업처분액"),
    ("투자수익", "투자수익 (h)", "기초부터 누계 투자수익"),
    ("운용수수료", "운용관리수수료 (i)", "기초부터 누계 운용관리수수료"),
    ("기말_퇴직연금", "기말잔액-퇴직연금", "작성기준일 현재 공정가치"),
    ("기말_퇴직신탁", "기말잔액-퇴직신탁", ""),
    ("기말_퇴직보험", "기말잔액-퇴직보험", ""),
    ("국민연금전환금", "국민연금전환금", "국민연금전환금이 있는 경우 입력"),
]
# 파싱 라벨 별칭(사용자가 양식을 바꿔도 잡히도록)
_FUNDING_ALIASES = {
    "기초잔액": ["기초잔액", "기초잔액A"],
    "입금액": ["입금수탁액B", "입금액", "입금수탁액"],
    "지급_퇴직": ["급여지급퇴직a", "퇴직a", "급여지급퇴직"],
    "지급_중간정산": ["급여지급중간정산b", "중간정산b", "급여지급중간정산"],
    "지급_DC전환": ["급여지급DC전환c", "DC전환c", "급여지급DC전환"],
    "관계사전입": ["관계사전입d", "관계사전입"],
    "관계사전출": ["관계사전출e", "관계사전출"],
    "사업결합": ["사업결합f", "사업결합"],
    "사업처분": ["사업처분g", "사업처분"],
    "투자수익": ["투자수익h", "투자수익"],
    "운용수수료": ["운용관리수수료i", "운용관리수수료", "운용수수료"],
    "기말_퇴직연금": ["기말잔액퇴직연금", "퇴직연금"],
    "기말_퇴직신탁": ["기말잔액퇴직신탁", "퇴직신탁"],
    "기말_퇴직보험": ["기말잔액퇴직보험", "퇴직보험"],
    "국민연금전환금": ["국민연금전환금"],
}


def build_funding_template(valuation_date: str = "") -> bytes:
    """사외적립자산 현황 양식(xlsx) — 사용자 제공 원본 시트 그대로.

    작성기준일 칸은 병합 셀이라 원본 그대로 두고, 사용자가 직접 입력한다.
    """
    return _to_bytes(_extract_sheet(_SHEET_FUNDING))


def parse_funding_upload(source: Union[str, bytes]) -> Dict:
    """사외적립자산 양식 업로드 → funding_status data_json dict.

    라벨(내용 열)을 정규화해 값(금액 열)을 읽는다. 작성기준일도 함께 반환.
    반환: {"_valuation_date": str|None, "공시방법": str, ...금액키: float}
    """
    ws = _open_first_sheet(source)
    grid = list(ws.iter_rows(values_only=True))
    alias_norm = {k: [_norm(a) for a in al] for k, al in _FUNDING_ALIASES.items()}

    out: Dict = {k: 0.0 for k, _l, _n in _FUNDING_ROWS}
    out["공시방법"] = ""
    out["_valuation_date"] = None

    for row in grid:
        if not row:
            continue
        cells = list(row)
        # 각 셀 라벨을 스캔, 같은 행에서 첫 숫자값을 값으로 사용
        labels = [_norm(c) for c in cells]
        # 값 후보: 이 행에서 숫자로 해석되는 셀
        numvals = [c for c in cells if isinstance(c, (int, float))]
        first_num = float(numvals[0]) if numvals else None

        for key, al in alias_norm.items():
            if any(a and a in lb for lb in labels for a in al):
                if first_num is not None:
                    out[key] = first_num
        # 작성기준일
        if any("작성기준일" in lb or "산출기준일" in lb for lb in labels):
            for c in cells:
                if isinstance(c, str) and len(c) >= 8 and any(ch.isdigit() for ch in c):
                    out["_valuation_date"] = c.strip()
        # 공시방법
        if any("공시방법" in lb for lb in labels):
            for c in cells:
                if isinstance(c, str) and ("공시" in c or c.strip() in ("①", "②", "③")):
                    out["공시방법"] = c.strip()
    return out


# ===========================================================================
# 2) 기타장기 포상제도
# ===========================================================================
def build_other_lt_template() -> bytes:
    """기타장기 포상제도 양식(xlsx) — 사용자 제공 원본 시트 그대로."""
    return _to_bytes(_extract_sheet(_SHEET_OTHER_LT))


def parse_other_lt_upload(source: Union[str, bytes]) -> Dict:
    """기타장기 양식 업로드 → {'지급내역': [...], '포상제도': [...]}.

    라벨행(헤더)을 찾아 그 아래 데이터행을 읽는다. 값이 있는 행만 반환.
    """
    ws = _open_first_sheet(source)
    grid = [list(r) for r in ws.iter_rows(values_only=True)]

    def _find_header(keys) -> Optional[int]:
        for ridx, row in enumerate(grid):
            norm = [_norm(c) for c in row]
            if all(any(k in nb for nb in norm) for k in keys):
                return ridx
        return None

    def _rows_after(hidx, ncols, stopwords=("합계", "ⓑ", "ⓒ", "포상제도")):
        out = []
        if hidx is None:
            return out
        for row in grid[hidx + 1:]:
            first = _norm(row[1]) if len(row) > 1 else ""
            if any(sw and sw in first for sw in map(_norm, stopwords)):
                break
            vals = [row[1 + k] if len(row) > 1 + k else None for k in range(ncols)]
            if any(v not in (None, "") for v in vals):
                out.append(vals)
        return out

    a_idx = _find_header(["근속년수", "근속자", "축하금"])
    b_idx = _find_header(["근속년수", "지급액", "유급휴가"])
    give = _rows_after(a_idx, 4)
    award = _rows_after(b_idx, 7)
    return {"지급내역": give, "포상제도": award}


# ===========================================================================
# 3) 명부확인용요약표
# ===========================================================================
# G열(자동산출) 항목 — (요약키, 라벨행 표시). 재직자명부에서 산출 가능한 항목만.
CENSUS_SUMMARY_AUTO_KEYS = [
    "재직_임원", "재직_직원", "재직_계약직", "재직_합계",
    "추계액_임원", "추계액_직원", "추계액_계약직", "추계액_합계",
    "중간정산자수", "중간정산금액",
]


def compute_census_summary(records) -> Dict[str, float]:
    """재직자명부 레코드 → 요약표 '자동산출' 값(G열용).

    재직자수(임원/직원/계약직/합계), 당년도 퇴직금추계액(구분별/합계),
    중간정산자수·중간정산금액 합계. (퇴직자·DC전환·관계사·사업결합은 다른 명부 필요 → 제외)
    """
    def _cls(r):
        v = getattr(r, "emp_class", None)
        return getattr(v, "value", v)

    n_exec = n_reg = n_con = 0
    a_exec = a_reg = a_con = 0.0
    interim_n = 0
    interim_amt = 0.0
    for r in records:
        cls = _cls(r)
        accr = float(getattr(r, "current_year_accrual", 0) or 0)
        if cls == "EXECUTIVE":
            n_exec += 1; a_exec += accr
        elif cls == "CONTRACT":
            n_con += 1; a_con += accr
        else:
            n_reg += 1; a_reg += accr
        if getattr(r, "interim_settlement_date", None) is not None:
            interim_n += 1
            interim_amt += float(getattr(r, "interim_settlement_amount", 0) or 0)
    return {
        "재직_임원": n_exec, "재직_직원": n_reg, "재직_계약직": n_con,
        "재직_합계": n_exec + n_reg + n_con,
        "추계액_임원": a_exec, "추계액_직원": a_reg, "추계액_계약직": a_con,
        "추계액_합계": a_exec + a_reg + a_con,
        "중간정산자수": interim_n, "중간정산금액": interim_amt,
    }


# 요약표 항목 행 정의 — (항목, 세부, 자동산출키|None, 비고)
_SUMMARY_ROWS = [
    ("평가대상", "재직자수-임원", "재직_임원", '"재직자 명부"의 임원 수'),
    ("", "재직자수-직원", "재직_직원", '"재직자 명부"의 직원 수'),
    ("", "재직자수-계약직", "재직_계약직", '"재직자 명부"의 계약직 수'),
    ("", "재직자수-합계", "재직_합계", "재직자 명부 합계와 일치"),
    ("", "퇴직자수-임원", None, '"퇴직자 명부" 필요 — 회사 입력'),
    ("", "퇴직자수-직원", None, ""),
    ("", "퇴직자수-계약직", None, ""),
    ("", "퇴직자수-합계", None, ""),
    ("", "중간정산자수", "중간정산자수", "당기 중간정산자 수"),
    ("", "DC전환자수", None, '"DC전환자 명부" 필요 — 회사 입력'),
    ("퇴직금추계액(원)", "임원", "추계액_임원", "당년도 퇴직금추계액-임원"),
    ("", "직원", "추계액_직원", "당년도 퇴직금추계액-직원"),
    ("", "계약직", "추계액_계약직", "당년도 퇴직금추계액-계약직"),
    ("", "합계", "추계액_합계", "재직자 명부 추계액 합계와 일치"),
    ("퇴직부채 감소(증가)액(원)", "급여지급-퇴직금 (a)", None, '"퇴직자 명부" 필요 — 회사 입력'),
    ("", "급여지급-중간정산 (b)", "중간정산금액", "중간정산금액 합계"),
    ("", "급여지급-DC전환 (c)", None, '"DC전환자 명부" 필요 — 회사 입력'),
    ("", "관계사전입 (d)", None, '"추가 명부" 필요 — 회사 입력'),
    ("", "관계사전출 (e)", None, ""),
    ("", "사업결합 (f)", None, ""),
    ("", "사업처분 (g)", None, ""),
]


def build_census_summary_template(summary: Optional[Dict[str, float]] = None,
                                  valuation_date: str = "") -> bytes:
    """명부확인용요약표 양식(xlsx) — 사용자 원본 시트에 G열(자동산출)을 채워 제공.

    summary: compute_census_summary() 결과. G열의 '등록부명에서 자동산출된 값' 칸에
    재직자명부 산출값(재직자수·추계액·중간정산)을 기입한다.
    """
    s = summary or {}
    wb = _extract_sheet(_SHEET_SUMMARY)
    ws = wb.worksheets[0]
    for cell, key in _SUMMARY_G_CELLS.items():
        val = s.get(key)
        if val is not None:
            ws[cell] = val
    if valuation_date:
        ws["F5"] = f"{valuation_date} 기준 요약표"   # 작성기준일 표시
    return _to_bytes(wb)


def parse_census_summary_upload(source: Union[str, bytes]) -> Dict:
    """명부확인용요약표 업로드 → 행별 {항목, 세부, 입력(F), 자동(G)} + 작성기준일."""
    ws = _open_first_sheet(source)
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    out: Dict = {"_valuation_date": None, "rows": []}
    cur_item = ""
    for row in grid:
        if not row:
            continue
        b = row[1] if len(row) > 1 else None       # B열 항목
        c = row[2] if len(row) > 2 else None       # C열 세부(재직자수 등)
        d = row[3] if len(row) > 3 else None       # D열 상세(임원/직원 등)
        f_val = row[5] if len(row) > 5 else None    # F열 회사 입력
        g_val = row[6] if len(row) > 6 else None    # G열 자동산출
        if b is not None and str(b).strip():
            cur_item = str(b).strip()
        # 작성기준일 행
        if c is not None and _norm(c) == "작성기준일":
            if f_val:
                out["_valuation_date"] = str(f_val).strip()
            continue
        label = " · ".join([str(x).strip() for x in (c, d) if x is not None and str(x).strip()])
        if not label:
            continue
        if _norm(label) in ("세부", "항목", "내용", "등록부명에서자동산출된값"):
            continue
        if f_val is None and g_val is None:
            continue
        out["rows"].append({"항목": cur_item, "세부": label,
                            "입력(회사)": f_val, "자동산출(명부)": g_val})
    return out
